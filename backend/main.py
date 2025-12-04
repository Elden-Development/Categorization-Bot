import io
import asyncio
import json
from fastapi import FastAPI, UploadFile, File, Body, Form, Depends, HTTPException, status, Request, Query, BackgroundTasks
import uuid
import threading
from dataclasses import dataclass, field
from typing import Dict, Any
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from pydantic import BaseModel, EmailStr, Field, field_validator
from dotenv import load_dotenv
import os
from typing import Optional, List, Literal
from datetime import datetime, date
from enum import Enum
# Rate limiting temporarily disabled - will be re-implemented with a different library
# from slowapi import Limiter, _rate_limit_exceeded_handler
# from slowapi.util import get_remote_address
# from slowapi.errors import RateLimitExceeded
from google import genai
from google.genai import types
from PyPDF2 import PdfReader, PdfWriter  # Install via pip install PyPDF2
from ml_categorization import get_ml_engine
from categories import get_all_categories, get_categories_by_parent, get_subcategories_for_category
from bank_statement_parser import BankStatementParser
from reconciliation_engine import ReconciliationEngine
from vendor_mapping import categorize_by_vendor, get_all_known_vendors, normalize_vendor_name

# Database imports
from sqlalchemy.orm import Session
from sqlalchemy import or_
from database import get_db, init_db, test_connection
from auth import get_current_user, get_optional_user, authenticate_user, create_access_token, hash_password
import crud
import models

# Load environment variables from .env file
load_dotenv()

# Helper function to safely navigate nested dicts/lists
def safe_get(data, *keys, default=None):
    """
    Safely navigate nested dicts/lists, returning default if any key fails.
    This prevents 'list indices must be integers' errors when data structure varies.
    """
    result = data
    for key in keys:
        if result is None:
            return default
        if isinstance(result, dict):
            result = result.get(key)
        elif isinstance(result, list) and isinstance(key, int) and 0 <= key < len(result):
            result = result[key]
        else:
            return default
    return result if result is not None else default


# Retry helper for API calls with rate limiting
async def retry_with_backoff(func, max_retries=3, initial_delay=2):
    """
    Retry an async function with exponential backoff.
    Specifically handles 429 RESOURCE_EXHAUSTED errors from Gemini API.

    Args:
        func: Async function to call (should be a coroutine or awaitable)
        max_retries: Maximum number of retry attempts
        initial_delay: Initial delay in seconds (doubles each retry)

    Returns:
        The result of the function call

    Raises:
        Exception: Re-raises the last exception if all retries fail
    """
    last_exception = None
    delay = initial_delay

    for attempt in range(max_retries + 1):
        try:
            # If func is a coroutine, await it; otherwise call it
            if asyncio.iscoroutine(func):
                return await func
            else:
                return await func()
        except Exception as e:
            last_exception = e
            error_str = str(e)

            # Check if it's a rate limit error (429)
            is_rate_limit = "429" in error_str or "RESOURCE_EXHAUSTED" in error_str

            if is_rate_limit and attempt < max_retries:
                print(f"Rate limit hit, retrying in {delay}s (attempt {attempt + 1}/{max_retries})")
                await asyncio.sleep(delay)
                delay *= 2  # Exponential backoff
            else:
                # Not a rate limit error or out of retries
                raise

    raise last_exception


def get_user_friendly_error(error: Exception) -> str:
    """Convert API errors to user-friendly messages."""
    error_str = str(error)

    if "429" in error_str or "RESOURCE_EXHAUSTED" in error_str:
        return "Service is temporarily busy. Please wait a moment and try again."
    elif "401" in error_str or "UNAUTHENTICATED" in error_str:
        return "API authentication error. Please contact support."
    elif "403" in error_str or "PERMISSION_DENIED" in error_str:
        return "API access denied. Please contact support."
    elif "500" in error_str or "INTERNAL" in error_str:
        return "The AI service is experiencing issues. Please try again later."
    elif "timeout" in error_str.lower():
        return "Request timed out. Please try again."
    else:
        return "An error occurred while processing your request. Please try again."


# ============================================================================
# GEMINI AI BANK STATEMENT PARSER
# ============================================================================

BANK_STATEMENT_PROMPT = """You are a bank statement parser. Analyze this bank statement PDF and extract ALL transactions.

For each transaction, extract:
1. date: The transaction date in YYYY-MM-DD format
2. description: The transaction description/memo/payee
3. amount: The transaction amount as a number (positive for deposits/credits, negative for withdrawals/debits)
4. type: Either "credit" (deposits, incoming money) or "debit" (withdrawals, outgoing money)
5. balance: The running balance after this transaction (if available)

Important parsing rules:
- Look for tabular data with columns like Date, Description, Debit, Credit, Balance
- If there are separate Debit and Credit columns, use the Debit value as negative amount and Credit value as positive
- Parse ALL transactions on ALL pages
- Ignore header rows, summary sections, and bank information
- Convert all dates to YYYY-MM-DD format
- Remove currency symbols from amounts, keep just the number
- If a transaction spans multiple lines, combine the description

Return a JSON object with this structure:
{
    "transactions": [
        {
            "date": "2024-01-15",
            "description": "PAYROLL DEPOSIT",
            "amount": 2500.00,
            "type": "credit",
            "balance": 5000.00
        },
        {
            "date": "2024-01-16",
            "description": "WALMART STORE #1234",
            "amount": -45.67,
            "type": "debit",
            "balance": 4954.33
        }
    ],
    "statement_info": {
        "bank_name": "Bank Name if found",
        "account_number": "Last 4 digits if found",
        "period_start": "Statement start date",
        "period_end": "Statement end date",
        "opening_balance": "Opening balance if found",
        "closing_balance": "Closing balance if found"
    }
}

Parse all transactions from the bank statement:"""


async def parse_bank_statement_with_gemini(file_content: bytes) -> List[Dict]:
    """
    Parse a PDF bank statement using Gemini AI.

    This is used as a fallback when the basic regex parser returns no results,
    which happens with tabular PDF formats.

    Parameters:
    file_content (bytes): PDF file content

    Returns:
    List[Dict]: List of extracted transactions
    """
    try:
        # Create a Gemini Part from the PDF bytes
        file_part = types.Part.from_bytes(
            data=file_content,
            mime_type="application/pdf"
        )

        # Call Gemini to extract transactions
        async def extract_transactions():
            return await asyncio.to_thread(
                client.models.generate_content,
                model="gemini-2.0-flash",
                contents=[BANK_STATEMENT_PROMPT, file_part],
                config={
                    "max_output_tokens": 40000,
                    "response_mime_type": "application/json"
                }
            )

        response = await retry_with_backoff(extract_transactions, max_retries=3, initial_delay=2)

        if not response or not response.text:
            print("Warning: Gemini returned empty response for bank statement")
            return []

        # Parse the JSON response
        result_text = response.text.strip()
        print(f"Gemini bank statement response length: {len(result_text)} chars")

        # Clean up potential markdown formatting
        if result_text.startswith("```json"):
            result_text = result_text[7:]
        if result_text.startswith("```"):
            result_text = result_text[3:]
        if result_text.endswith("```"):
            result_text = result_text[:-3]
        result_text = result_text.strip()

        parsed_data = json.loads(result_text)

        transactions = parsed_data.get("transactions", [])
        statement_info = parsed_data.get("statement_info", {})

        print(f"Gemini extracted {len(transactions)} transactions from bank statement")

        # Normalize transactions to expected format
        normalized_transactions = []
        for idx, tx in enumerate(transactions):
            normalized = {
                "transaction_id": f"gemini_tx_{idx}",
                "date": tx.get("date"),
                "description": tx.get("description", "").strip(),
                "amount": tx.get("amount"),
                "type": tx.get("type", "debit" if tx.get("amount", 0) < 0 else "credit"),
                "source": "gemini_ai"  # Mark source for debugging
            }

            # Include balance if available
            if tx.get("balance") is not None:
                normalized["balance"] = tx.get("balance")

            # Only include transactions with required fields
            if normalized["date"] and normalized["amount"] is not None:
                normalized_transactions.append(normalized)

        return normalized_transactions

    except json.JSONDecodeError as e:
        print(f"Error parsing Gemini JSON response: {str(e)}")
        return []
    except Exception as e:
        print(f"Error parsing bank statement with Gemini: {str(e)}")
        return []


# ============================================================================
# BATCH JOB TRACKING SYSTEM
# ============================================================================

@dataclass
class BatchJob:
    """Represents a batch processing job"""
    job_id: str
    user_id: int
    statement_id: int
    status: str  # 'pending', 'processing', 'completed', 'failed'
    total_transactions: int = 0
    processed_count: int = 0
    failed_count: int = 0
    high_confidence_count: int = 0
    low_confidence_count: int = 0
    current_transaction: str = ""
    progress_percent: float = 0.0
    started_at: datetime = field(default_factory=datetime.now)
    completed_at: Optional[datetime] = None
    error_message: Optional[str] = None
    results: list = field(default_factory=list)
    category_counts: dict = field(default_factory=dict)


class BatchJobTracker:
    """Thread-safe batch job tracker"""

    def __init__(self):
        self._jobs: Dict[str, BatchJob] = {}
        self._lock = threading.Lock()
        self._max_jobs = 1000  # Max jobs to keep in memory

    def create_job(self, user_id: int, statement_id: int, total_transactions: int) -> str:
        """Create a new batch job and return job_id"""
        job_id = str(uuid.uuid4())

        with self._lock:
            # Clean up old jobs if we have too many
            if len(self._jobs) >= self._max_jobs:
                self._cleanup_old_jobs()

            self._jobs[job_id] = BatchJob(
                job_id=job_id,
                user_id=user_id,
                statement_id=statement_id,
                status="pending",
                total_transactions=total_transactions
            )

        return job_id

    def start_job(self, job_id: str):
        """Mark job as processing"""
        with self._lock:
            if job_id in self._jobs:
                self._jobs[job_id].status = "processing"
                self._jobs[job_id].started_at = datetime.now()

    def update_progress(self, job_id: str, processed: int, current_description: str = "",
                       high_conf: int = 0, low_conf: int = 0, failed: int = 0):
        """Update job progress"""
        with self._lock:
            if job_id in self._jobs:
                job = self._jobs[job_id]
                job.processed_count = processed
                job.current_transaction = current_description
                job.high_confidence_count = high_conf
                job.low_confidence_count = low_conf
                job.failed_count = failed
                if job.total_transactions > 0:
                    job.progress_percent = round((processed / job.total_transactions) * 100, 1)

    def add_result(self, job_id: str, result: dict):
        """Add a categorization result to the job"""
        with self._lock:
            if job_id in self._jobs:
                self._jobs[job_id].results.append(result)

    def update_category_count(self, job_id: str, category: str):
        """Update category distribution"""
        with self._lock:
            if job_id in self._jobs:
                counts = self._jobs[job_id].category_counts
                counts[category] = counts.get(category, 0) + 1

    def complete_job(self, job_id: str, success: bool = True, error_message: str = None):
        """Mark job as completed or failed"""
        with self._lock:
            if job_id in self._jobs:
                job = self._jobs[job_id]
                job.status = "completed" if success else "failed"
                job.completed_at = datetime.now()
                job.progress_percent = 100.0 if success else job.progress_percent
                job.error_message = error_message

    def get_job(self, job_id: str) -> Optional[BatchJob]:
        """Get job by ID"""
        with self._lock:
            return self._jobs.get(job_id)

    def get_user_jobs(self, user_id: int, limit: int = 10) -> list:
        """Get recent jobs for a user"""
        with self._lock:
            user_jobs = [j for j in self._jobs.values() if j.user_id == user_id]
            # Sort by started_at descending
            user_jobs.sort(key=lambda x: x.started_at, reverse=True)
            return user_jobs[:limit]

    def _cleanup_old_jobs(self):
        """Remove oldest completed jobs to free memory"""
        completed_jobs = [(jid, j) for jid, j in self._jobs.items()
                         if j.status in ('completed', 'failed')]
        # Sort by completion time
        completed_jobs.sort(key=lambda x: x[1].completed_at or x[1].started_at)
        # Remove oldest half
        for jid, _ in completed_jobs[:len(completed_jobs)//2]:
            del self._jobs[jid]


# Global batch job tracker instance
batch_job_tracker = BatchJobTracker()


app = FastAPI(title="Categorization Bot API", version="1.0.0")

# Initialize rate limiter - TEMPORARILY DISABLED due to conflicts with Pydantic request models
# TODO: Re-implement rate limiting with a different approach that doesn't conflict with Pydantic
# limiter = Limiter(key_func=get_remote_address, default_limits=["200/hour", "50/minute"])
# app.state.limiter = limiter
# app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# Enable CORS - configured via environment variable for security
# Default to localhost for development. In production, set CORS_ORIGINS in .env
cors_origins_str = os.getenv("CORS_ORIGINS", "http://localhost:3000,http://localhost:3001")
cors_origins = [origin.strip() for origin in cors_origins_str.split(",")]

# Add Railway domains automatically
cors_origins.extend([
    "https://frontend-production-e172.up.railway.app",
    "https://backend-production-3336.up.railway.app",
])
# Remove duplicates
cors_origins = list(set(cors_origins))

print(f"[CORS] Allowed origins: {cors_origins}")

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Database initialization on startup
@app.on_event("startup")
async def startup_event():
    """Initialize database on application startup"""
    try:
        print("Testing database connection...")
        if test_connection():
            print("Initializing database tables...")
            init_db()
            print("[OK] Database initialized successfully!")
        else:
            print("[WARNING] Database connection failed. Running without persistence.")
            print("  To enable database features:")
            print("  1. Install PostgreSQL")
            print("  2. Create database: categorization_bot")
            print("  3. Set DATABASE_URL in .env file")
    except Exception as e:
        print(f"[WARNING] Database initialization failed: {e}")
        print("  Application will run without data persistence.")

# Gemini API key loaded from environment variables
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
PINECONE_API_KEY = os.getenv("PINECONE_API_KEY")
client = genai.Client(api_key=GEMINI_API_KEY)


# =============================================================================
# HEALTH CHECK ENDPOINTS
# =============================================================================

@app.get("/")
async def root():
    """Root endpoint - basic API info"""
    return {
        "name": "Categorization Bot API",
        "version": "1.0.0",
        "status": "running",
        "docs": "/docs"
    }


@app.get("/health")
async def health_check(db: Session = Depends(get_db)):
    """
    Health check endpoint for monitoring and deployment verification.

    Returns status of all critical services:
    - API server
    - Database connection
    - Gemini API
    - Pinecone (ML engine)
    """
    health_status = {
        "status": "healthy",
        "timestamp": datetime.utcnow().isoformat(),
        "services": {
            "api": {"status": "up"},
            "database": {"status": "unknown"},
            "gemini": {"status": "unknown"},
            "pinecone": {"status": "unknown"}
        }
    }

    # Check database
    try:
        if test_connection():
            health_status["services"]["database"] = {"status": "up"}
        else:
            health_status["services"]["database"] = {"status": "down", "error": "Connection failed"}
            health_status["status"] = "degraded"
    except Exception as e:
        health_status["services"]["database"] = {"status": "down", "error": str(e)}
        health_status["status"] = "degraded"

    # Check Gemini API
    try:
        if GEMINI_API_KEY:
            # Quick test - just verify client is initialized
            health_status["services"]["gemini"] = {"status": "configured"}
        else:
            health_status["services"]["gemini"] = {"status": "not_configured", "error": "API key missing"}
            health_status["status"] = "degraded"
    except Exception as e:
        health_status["services"]["gemini"] = {"status": "error", "error": str(e)}
        health_status["status"] = "degraded"

    # Check Pinecone
    try:
        if PINECONE_API_KEY:
            health_status["services"]["pinecone"] = {"status": "configured"}
        else:
            health_status["services"]["pinecone"] = {"status": "not_configured", "note": "ML features disabled"}
    except Exception as e:
        health_status["services"]["pinecone"] = {"status": "error", "error": str(e)}

    return health_status


@app.get("/health/quick")
async def health_check_quick():
    """Quick health check - just returns OK if server is running"""
    return {"status": "ok"}

# Initialize ML Engine (lazy initialization on first use)
ml_engine = None

def get_ml_categorization_engine():
    """Get or initialize the ML categorization engine."""
    global ml_engine
    if ml_engine is None:
        if not PINECONE_API_KEY:
            raise ValueError("PINECONE_API_KEY not found in environment variables")
        ml_engine = get_ml_engine(
            pinecone_api_key=PINECONE_API_KEY,
            gemini_api_key=GEMINI_API_KEY
        )
    return ml_engine

# Schema validation
class DocumentSchema(str, Enum):
    """Valid document schema types"""
    GENERIC = "generic"
    FORM_1040 = "1040"
    FORM_2848 = "2848"
    FORM_8821 = "8821"
    FORM_941 = "941"
    PAYROLL = "payroll"

# File validation configuration
MAX_FILE_SIZE = 25 * 1024 * 1024  # 25 MB in bytes
ALLOWED_MIME_TYPES = {
    "application/pdf",
    "image/png",
    "image/jpeg",
    "image/jpg",
    "image/gif",
    "image/bmp",
    "image/tiff",
    "text/csv",
    "application/csv",
    "text/plain"  # CSV files sometimes reported as plain text
}
ALLOWED_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".tif", ".csv"}

async def validate_file_upload(file: UploadFile) -> None:
    """
    Validate uploaded file for security and size constraints.

    Raises HTTPException if validation fails.
    """
    # Check file extension
    if file.filename:
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail=f"File type not allowed. Allowed types: {', '.join(ALLOWED_EXTENSIONS)}"
            )

    # Check MIME type
    if file.content_type not in ALLOWED_MIME_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"MIME type not allowed. Allowed types: {', '.join(ALLOWED_MIME_TYPES)}"
        )

    # Check file size by reading in chunks to avoid loading entire file
    file_size = 0
    chunk_size = 1024 * 1024  # 1 MB chunks

    # Read file to check size
    content = await file.read()
    file_size = len(content)

    # Reset file pointer for later processing
    await file.seek(0)

    if file_size > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"File too large. Maximum size: {MAX_FILE_SIZE / (1024*1024):.0f}MB"
        )

    if file_size == 0:
        raise HTTPException(
            status_code=400,
            detail="Empty file not allowed"
        )

# Step 1: Raw extraction prompt remains unchanged.
RAW_PROMPT = "List every single thing exactly as it appears on the document, each column and row, in full"

# Helper function to load a schema file
def load_schema(schema_id):
    """
    Load the specified JSON schema from file
    
    Parameters:
    schema_id (str): Identifier for the schema to load
    
    Returns:
    dict: The loaded schema or None if not found
    """
    schema_mapping = {
        "1040": "1040.json",
        "2848": "2848.json",
        "8821": "8821.json",
        "941": "941.json",
        "payroll": "payroll.json",
        "generic": None  # Generic schema doesn't need a specific file
    }
    
    if schema_id not in schema_mapping:
        return None
        
    filename = schema_mapping.get(schema_id)
    if filename is None:
        return None
    
    # Get the directory where the script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # Go up one level to the project root
    project_root = os.path.dirname(script_dir)
    # Construct absolute path to the schema file
    schema_path = os.path.join(project_root, filename)
    
    try:
        with open(schema_path, 'r') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"Error loading schema {schema_id} from {schema_path}: {e}")
        return None

# Function to generate a schema-specific prompt
def generate_schema_prompt(schema_id, extracted_text):
    """
    Generate a prompt for extraction based on the selected schema
    
    Parameters:
    schema_id (str): Identifier for the schema to use
    extracted_text (str): Raw text extracted from the document
    
    Returns:
    str: The prompt to use for extraction
    """
    # Load the requested schema
    schema = load_schema(schema_id)
    
    # Generic document schema for fallback
    GENERIC_SCHEMA = """
    {
      "documentMetadata": {
        "documentID": "Unique identifier for the document",
        "documentType": "Type/category (e.g., Invoice, Receipt, BankStatement, MerchantStatement, CreditMemo, PurchaseOrder, etc.)",
        "documentNumber": "Official reference number from the document",
        "documentDate": "YYYY-MM-DD (date issued)",
        "uploadDate": "YYYY-MM-DDTHH:MM:SSZ (date/time uploaded)",
        "source": {
          "name": "Name of the issuer (vendor, bank, customer, etc.)",
          "type": "Category (Vendor, Customer, Bank, etc.)",
          "contact": {
            "address": "Full address if available",
            "phone": "Contact phone number",
            "email": "Contact email address"
          }
        },
        "fileInfo": {
          "fileName": "Original file name",
          "fileType": "Format (e.g., PDF, JPEG)",
          "pageCount": "Number of pages in the document",
          "OCRProcessed": "Boolean flag indicating whether OCR has been applied"
        }
      },
      "financialData": {
        "currency": "Currency code (e.g., USD, EUR)",
        "subtotal": "Amount before any discounts or additional fees",
        "taxAmount": "Total tax applied (if any) – can be zero or omitted if not applicable",
        "discount": "Any discounts or adjustments applied",
        "totalAmount": "Final total monetary value on the document",
        "paymentStatus": "Status (e.g., Paid, Unpaid, Pending)",
        "paymentTerms": "Terms such as 'Net 30'",
        "dueDate": "YYYY-MM-DD (if a due date is specified)"
      },
      "lineItems": [
        {
          "itemID": "Unique identifier for the line item or transaction",
          "description": "Short description of the product, service, or transaction",
          "quantity": "Number of units or hours (if invoicing for services)",
          "unit": "Unit of measure (e.g., hours, pieces, kg, etc.)",
          "unitPrice": "Price per unit or hourly rate",
          "totalPrice": "Calculated total for the line item",
          "tax": "Tax amount applicable to this line item (if any)",
          "transactionType": "For bank/merchant statements (e.g., Debit, Credit)",
          "balance": "Running balance after the transaction (if applicable)",
          "category": "Optional field to classify the line item (e.g., Service, Product)"
        }
      ],
      "partyInformation": {
        "vendor": {
          "name": "Vendor/Supplier name",
          "address": "Vendor address",
          "contact": "Vendor contact details (phone/email)",
          "taxID": "Tax identifier if available (optional)"
        },
        "customer": {
          "name": "Customer name",
          "address": "Customer address",
          "contact": "Customer contact details",
          "customerID": "Internal customer identifier"
        },
        "bankDetails": {
          "bankName": "Name of the bank",
          "accountNumber": "Bank account number",
          "routingNumber": "Routing or sort code"
        }
      },
      "paymentInformation": {
        "paymentMethod": "Method (e.g., Cash, Credit Card, EFT, Direct Deposit)",
        "transactionID": "Identifier for electronically processed payments",
        "paidDate": "YYYY-MM-DD (date when payment was made)",
        "bankDetails": {
          "bankName": "If relevant, bank name for the transaction",
          "accountNumber": "Account number associated with payment",
          "routingNumber": "Routing number for bank transfers"
        }
      },
      "fixedAssetData": {
        "assetID": "Unique asset identifier",
        "description": "Description of the fixed asset or inventory item",
        "acquisitionDate": "YYYY-MM-DD (date of purchase/acquisition)",
        "purchasePrice": "Cost of acquiring the asset",
        "depreciationMethod": "Method used (e.g., Straight-line, Declining Balance)",
        "currentValue": "Current book value of the asset",
        "location": "Physical location of the asset (if applicable)"
      },
      "additionalData": {
        "notes": "Any annotations or internal comments",
        "attachments": [
          "Link(s) or reference(s) to supplementary files, if applicable"
        ],
        "referenceNumbers": [
          "Other relevant reference numbers (e.g., purchase order numbers, shipping numbers)"
        ],
        "auditTrail": [
          {
            "action": "Description of the processing step (e.g., 'uploaded', 'OCR extracted')",
            "timestamp": "YYYY-MM-DDTHH:MM:SSZ",
            "user": "User or system that performed the action"
          }
        ]
      }
    }
    """
    
    # Base prompt template
    prompt_template = """
    Format the following extracted text into a JSON object that strictly follows the schema below. Ensure that the output is valid JSON with no additional commentary.
    
    Schema:
    {schema}
    
    Extracted Text:
    {extracted_text}
    """
    
    # If we have a specific schema, use it; otherwise, use the generic one
    if schema:
        schema_json = json.dumps(schema, indent=2)
        
        # For tax forms, add some specific instructions
        if schema_id in ["1040", "2848", "8821", "941"]:
            schema_prompt = f"""
            Format the following extracted text from an IRS tax form into a JSON object that strictly follows the schema below.
            This is specifically for IRS Form {schema_id}. Pay special attention to the form fields, numbers, checkboxes, and taxpayer information.
            
            Schema:
            {schema_json}
            
            Extracted Text:
            {extracted_text}
            """
        # For payroll data
        elif schema_id == "payroll":
            schema_prompt = f"""
            Format the following extracted text from a payroll document into a JSON object that strictly follows the schema below.
            This is specifically for payroll data. Pay special attention to employee details, earnings, deductions, and tax information.
            
            Schema:
            {schema_json}
            
            Extracted Text:
            {extracted_text}
            """
        else:
            schema_prompt = prompt_template.format(schema=schema_json, extracted_text=extracted_text)
    else:
        # Use the generic schema as fallback
        schema_prompt = prompt_template.format(schema=GENERIC_SCHEMA, extracted_text=extracted_text)
    
    return schema_prompt

def detect_document_type(json_data):
    """
    Detect the document type from the JSON data to apply appropriate verification rules.
    
    Parameters:
    json_data (dict): The structured JSON data extracted from the document
    
    Returns:
    str: Document type - 'invoice', 'receipt', 'payment_processing', 'bank_statement', or 'other'
    """
    # Get the document type from metadata if available
    doc_type_raw = safe_get(json_data, "documentMetadata", "documentType", default="")
    doc_type = str(doc_type_raw).lower() if doc_type_raw else ""
    
    # Check for specific document type indicators
    if doc_type in ["merchantstatement", "merchant_statement", "payment_processing_statement", 
                    "processor_statement", "acquirer_statement"]:
        return "payment_processing"
    
    # Check for bank statement indicators
    if doc_type in ["bankstatement", "bank_statement", "account_statement"]:
        return "bank_statement"
    
    # Check for invoice indicators
    if doc_type in ["invoice", "bill"]:
        return "invoice"
    
    # Check for receipt indicators
    if doc_type in ["receipt", "sales_receipt"]:
        return "receipt"
    
    # If no explicit type, analyze the content to determine type
    
    # Check for payment processing statement indicators
    if any(keyword in str(json_data).lower() for keyword in 
           ["interchange", "merchant id", "card summary", "chargebacks", 
            "settlement", "processor", "acquirer", "mastercard", "visa fees"]):
        return "payment_processing"
    
    # Check for card transaction indicators (high volume of transactions)
    line_items = json_data.get("lineItems", [])
    if len(line_items) > 20 and any("card" in str(item).lower() for item in line_items):
        return "payment_processing"
    
    # Check for special fields that indicate payment processing
    if "credits" in str(json_data).lower() and "sales" in str(json_data).lower() and "settlement" in str(json_data).lower():
        return "payment_processing"
    
    # Check for indicators of a bank statement
    if any(keyword in str(json_data).lower() for keyword in 
           ["account number", "routing number", "beginning balance", "ending balance", 
            "deposits", "withdrawals", "account summary"]):
        return "bank_statement"
    
    # Check for invoice indicators (if not already detected)
    if "invoice" in str(json_data).lower() or "bill to" in str(json_data).lower():
        return "invoice"
    
    # Default to invoice verification rules if we can't determine
    return "invoice"

async def verify_extraction(json_data):
    """
    Verify the mathematical accuracy of the extracted data by focusing on 
    calculation discrepancies rather than trivial formatting differences.
    
    Parameters:
    json_data (dict): The structured JSON data extracted from the document
    
    Returns:
    dict: Original JSON with added extraction verification results
    """
    try:
        # First, detect document type to apply appropriate verification rules
        document_type = detect_document_type(json_data)
        
        # Update the prompt to focus on mathematical verification
        verification_prompt = f"""
        Analyze this financial document data to verify MATHEMATICAL ACCURACY only.
        
        DOCUMENT TYPE: {document_type.upper()}
        
        CRITICAL INSTRUCTIONS:
        1. IGNORE all formatting differences (e.g., "90" vs 90, spaces, character encoding)
        2. IGNORE data type differences (string vs number) - only care about the numeric value
        3. FOCUS ONLY on verification of financial calculations:
           - For invoices: Check if quantity × price = line totals
           - For all documents: Verify line items sum up to stated subtotals or totals
           - Check if subtotal + tax = total amount (where applicable)
           - For statements: Verify opening balance + transactions = closing balance
         
        4. Mathematical rules that should be checked:
           - Line items: quantity × unit price = line total
           - Document totals: sum of line totals = subtotal
           - Tax calculation: subtotal × tax rate = tax amount
           - Final total: subtotal + tax + fees - discounts = total amount
        
        Financial Data:
        {json.dumps(json_data, indent=2)}
        
        Return your verification as JSON with this strict structure:
        {{
          "extractionVerified": Boolean (true if calculations are accurate, false if calculation errors exist),
          "discrepancies": [
            {{
              "type": "String (Line Total, Subtotal, Tax Calculation, etc.)",
              "location": "String (reference to where in the document this calculation occurs)",
              "expectedValue": "Number (what the calculation should produce)",
              "extractedValue": "Number (what was extracted in the document)",
              "likelyCorrectValue": "Number (the most likely correct value)",
              "formula": "String (the formula used for this calculation)",
              "confidence": "String (High, Medium, Low)"
            }}
          ],
          "summary": "String (brief explanation of calculation verification results)"
        }}
        
        ONLY include discrepancies that are TRUE CALCULATION ERRORS where numbers don't add up correctly.
        DO NOT flag differences in formatting, string representations, or character encoding.
        """
        
        # Make the API call to Gemini
        verification_response = await asyncio.to_thread(
            client.models.generate_content,
            model="gemini-2.0-flash",
            contents=[verification_prompt],
            config={
                "max_output_tokens": 4000,
                "response_mime_type": "application/json"
            }
        )
        
        # Extract verification results
        try:
            verification_results = json.loads(verification_response.text)
            
            # Keep only significant calculation discrepancies
            if "discrepancies" in verification_results and len(verification_results["discrepancies"]) > 0:
                # Filter out any discrepancies where the numeric values are actually the same
                significant_issues = []
                for d in verification_results["discrepancies"]:
                    # Try to convert both values to floats for comparison
                    try:
                        expected = float(str(d.get("expectedValue", "0")).replace(",", ""))
                        actual = float(str(d.get("extractedValue", "0")).replace(",", ""))
                        
                        # Only keep discrepancies where values are numerically different
                        if abs(expected - actual) > 0.01:  # Allow for small rounding differences
                            significant_issues.append(d)
                    except:
                        # If we can't convert to float, keep the discrepancy
                        significant_issues.append(d)
                
                verification_results["discrepancies"] = significant_issues
                
                # Update the verification status based on filtered discrepancies
                verification_results["extractionVerified"] = len(significant_issues) == 0
                
                # Update summary if needed
                if len(significant_issues) == 0 and not verification_results["extractionVerified"]:
                    verification_results["summary"] = "No significant calculation discrepancies found after filtering."
                    verification_results["extractionVerified"] = True
            
            # Add verification results to the original JSON
            json_data["extractionVerification"] = verification_results
            
        except json.JSONDecodeError as e:
            # If the response isn't valid JSON, create a simple error structure
            json_data["extractionVerification"] = {
                "extractionVerified": False,
                "discrepancies": [],
                "summary": f"Error parsing verification results: {str(e)}",
                "rawResponse": verification_response.text
            }
        
        return json_data
        
    except Exception as e:
        # If verification fails, add error information to the JSON
        json_data["extractionVerification"] = {
            "extractionVerified": False,
            "discrepancies": [],
            "summary": f"Error during extraction verification: {str(e)}"
        }
        return json_data

async def process_page(page, schema="generic"):
    """
    Process a single PDF page and extract structured data.
    Includes retry logic for rate limits and proper error handling.
    """
    # Write the individual page to a BytesIO stream.
    pdf_writer = PdfWriter()
    pdf_writer.add_page(page)
    page_stream = io.BytesIO()
    pdf_writer.write(page_stream)
    page_stream.seek(0)

    page_bytes = page_stream.getvalue()
    print(f"Processing PDF page, size: {len(page_bytes)} bytes")

    # Create a Gemini Part from the page bytes.
    file_part = types.Part.from_bytes(
        data=page_bytes,
        mime_type="application/pdf"
    )

    # Step 1: Extract raw text from the page (with retry for rate limits)
    async def extract_raw_text():
        return await asyncio.to_thread(
            client.models.generate_content,
            model="gemini-2.0-flash",
            contents=[RAW_PROMPT, file_part],
            config={
                "max_output_tokens": 40000,
                "response_mime_type": "text/plain"
            }
        )

    try:
        raw_response = await retry_with_backoff(extract_raw_text, max_retries=3, initial_delay=2)
    except Exception as e:
        print(f"Error extracting raw text from PDF page: {str(e)}")
        # Return error JSON that merge_page_results can handle
        return json.dumps({"error": f"Failed to extract text: {get_user_friendly_error(e)}"})

    raw_text = raw_response.text if raw_response else ""
    print(f"Raw text extracted, length: {len(raw_text) if raw_text else 0} chars")

    # Check if raw text extraction returned empty
    if not raw_text or raw_text.strip() == "":
        print("Warning: Gemini returned empty text for PDF page")
        return json.dumps({"error": "No text could be extracted from this page"})

    # Step 2: Convert the raw text into structured JSON using the schema-specific prompt
    json_prompt = generate_schema_prompt(schema, raw_text)

    async def convert_to_json():
        return await asyncio.to_thread(
            client.models.generate_content,
            model="gemini-2.0-flash",
            contents=[json_prompt],
            config={
                "max_output_tokens": 40000,
                "response_mime_type": "application/json"
            }
        )

    try:
        json_response = await retry_with_backoff(convert_to_json, max_retries=3, initial_delay=2)
    except Exception as e:
        print(f"Error converting to JSON: {str(e)}")
        return json.dumps({"error": f"Failed to structure data: {get_user_friendly_error(e)}"})

    result = json_response.text if json_response else ""
    print(f"JSON response generated, length: {len(result) if result else 0} chars")

    if not result or result.strip() == "":
        print("Warning: Gemini returned empty JSON response")
        return json.dumps({"error": "Failed to generate structured data from page"})

    # Return the JSON response to be merged later
    return result

def deep_merge(base, addition):
    """
    Recursively merge two dictionaries.
    - Lists are concatenated
    - Dictionaries are merged recursively
    - For other values, non-null values are preferred over null values
    - For other cases where both values are non-null, the first occurrence (base) is kept
    
    Parameters:
    base (dict): The base dictionary to merge into
    addition (dict): The dictionary to merge from
    
    Returns:
    dict: The merged dictionary
    """
    # Create a copy to avoid modifying the original
    result = base.copy()
    
    for key, value in addition.items():
        # If key not in result, just add it
        if key not in result:
            result[key] = value
        else:
            # If both are lists, extend the base list
            if isinstance(result[key], list) and isinstance(value, list):
                result[key].extend(value)
            
            # If both are dictionaries, merge them recursively
            elif isinstance(result[key], dict) and isinstance(value, dict):
                result[key] = deep_merge(result[key], value)
            
            # For boolean values, use logical OR (True if either is True)
            elif isinstance(result[key], bool) and isinstance(value, bool):
                result[key] = result[key] or value
            
            # For other values, prefer non-null values over null/empty ones
            elif result[key] is None and value is not None:
                result[key] = value
            # If both values exist and neither is None, keep the base value (first page)
    
    return result

def merge_page_results(page_results):
    """
    Merge JSON results from multiple pages.
    For document-level fields, we assume they are the same across pages and only keep the first occurrence.
    For list fields, we concatenate them, regardless of where they appear in the JSON structure.
    """
    merged = None
    failed_pages = 0
    error_messages = []

    # Process each page
    for idx, result in enumerate(page_results):
        # Handle None or empty results
        if result is None or result == "":
            print(f"Warning: Page {idx + 1} returned empty result")
            failed_pages += 1
            error_messages.append(f"Page {idx + 1}: Empty result")
            continue

        try:
            data = json.loads(result)
        except json.JSONDecodeError as e:
            print(f"Warning: Page {idx + 1} JSON decode error: {e}")
            failed_pages += 1
            error_messages.append(f"Page {idx + 1}: Invalid JSON")
            continue

        # Check if the page returned an error object
        if isinstance(data, dict) and "error" in data:
            print(f"Warning: Page {idx + 1} returned error: {data.get('error')}")
            failed_pages += 1
            error_messages.append(f"Page {idx + 1}: {data.get('error')}")
            continue

        if merged is None:
            merged = data
        else:
            # Use deep merge to recursively combine the data
            merged = deep_merge(merged, data)

    # If all pages failed, return an error structure instead of None
    if merged is None:
        print(f"Error: All {len(page_results)} pages failed to parse")
        # Provide more specific error message
        if error_messages:
            detail = "; ".join(error_messages[:3])  # Show first 3 errors
            if len(error_messages) > 3:
                detail += f" (and {len(error_messages) - 3} more)"
        else:
            detail = "The document may be scanned/image-based or contain unreadable content."

        return {
            "error": "Failed to extract data from document",
            "detail": detail,
            "failed_pages": failed_pages,
            "total_pages": len(page_results)
        }

    # Remove any extractionVerification fields - we'll perform a new verification on the complete document
    if merged and "extractionVerification" in merged:
        del merged["extractionVerification"]

    return merged

@app.post("/process-pdf")
async def process_file(
    file: UploadFile = File(...),
    schema: str = Form("generic"),
    current_user: models.User = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """
    Process a document (PDF or image) and extract structured data.

    If user is authenticated, document will be saved to database for history.
    """
    # Validate schema parameter
    valid_schemas = [s.value for s in DocumentSchema]
    if schema not in valid_schemas:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid schema. Must be one of: {', '.join(valid_schemas)}"
        )

    # Validate file upload
    await validate_file_upload(file)

    file_content = await file.read()

    # Generate unique document ID
    import uuid
    document_id = str(uuid.uuid4())

    # Save document to database if user is authenticated
    db_document = None
    if current_user:
        try:
            db_document = crud.create_document(
                db=db,
                user_id=current_user.id,
                document_id=document_id,
                file_name=file.filename,
                file_type=file.content_type,
                file_size=len(file_content),
                schema_type=schema
            )

            # Update status to processing
            crud.update_document_status(
                db=db,
                document_id=document_id,
                user_id=current_user.id,
                status="processing",
                progress=10
            )

            # Log activity
            crud.log_activity(
                db=db,
                user_id=current_user.id,
                action="document_uploaded",
                entity_type="document",
                entity_id=db_document.id,
                details={"file_name": file.filename, "schema": schema}
            )
        except Exception as e:
            print(f"Warning: Failed to save document to database: {e}")

    try:
        if file.content_type == "application/pdf":
            pdf_reader = PdfReader(io.BytesIO(file_content))
            # Process each page concurrently using the per-page processing function.
            tasks = [process_page(page, schema) for page in pdf_reader.pages]
            page_results = await asyncio.gather(*tasks)

            # Merge the JSON results from each page.
            merged_result = merge_page_results(page_results)

            # Check if merge_page_results returned an error
            if merged_result and isinstance(merged_result, dict) and "error" in merged_result:
                error_detail = merged_result.get("detail", "Failed to extract data from document")
                # Update document status to error if user is authenticated
                if current_user and db_document:
                    try:
                        crud.update_document_status(
                            db=db,
                            document_id=document_id,
                            user_id=current_user.id,
                            status="error",
                            error_message=error_detail
                        )
                    except:
                        pass
                raise HTTPException(status_code=422, detail=error_detail)

            # Perform extraction verification on the complete document
            final_result = await verify_extraction(merged_result)

            combined_response_text = json.dumps(final_result, indent=2)
        else:
            # For non-PDF files, process with schema selection
            file_part = types.Part.from_bytes(
                data=file_content,
                mime_type=file.content_type
            )
            raw_response = await asyncio.to_thread(
                client.models.generate_content,
                model="gemini-2.0-flash",
                contents=[RAW_PROMPT, file_part],
                config={
                    "max_output_tokens": 40000,
                    "response_mime_type": "text/plain"
                }
            )
            raw_text = raw_response.text

            # Check if raw text extraction failed
            if not raw_text or raw_text.strip() == "":
                error_detail = "Failed to extract text from image. The image may be unreadable or contain no text."
                if current_user and db_document:
                    try:
                        crud.update_document_status(
                            db=db,
                            document_id=document_id,
                            user_id=current_user.id,
                            status="error",
                            error_message=error_detail
                        )
                    except:
                        pass
                raise HTTPException(status_code=422, detail=error_detail)

            # Use schema-specific prompt template
            json_prompt = generate_schema_prompt(schema, raw_text)

            json_response = await asyncio.to_thread(
                client.models.generate_content,
                model="gemini-2.0-flash",
                contents=[json_prompt],
                config={
                    "max_output_tokens": 40000,
                    "response_mime_type": "application/json"
                }
            )

            # Validate Gemini response
            if not json_response or not json_response.text or json_response.text.strip() == "":
                error_detail = "AI model returned empty response. Please try again or use a different document."
                if current_user and db_document:
                    try:
                        crud.update_document_status(
                            db=db,
                            document_id=document_id,
                            user_id=current_user.id,
                            status="error",
                            error_message=error_detail
                        )
                    except:
                        pass
                raise HTTPException(status_code=422, detail=error_detail)

            # For non-PDF files (single page), add extraction verification
            try:
                json_data = json.loads(json_response.text)
                final_json = await verify_extraction(json_data)
                combined_response_text = json.dumps(final_json, indent=2)
            except json.JSONDecodeError as e:
                error_detail = f"Failed to parse AI response as JSON: {str(e)}"
                if current_user and db_document:
                    try:
                        crud.update_document_status(
                            db=db,
                            document_id=document_id,
                            user_id=current_user.id,
                            status="error",
                            error_message=error_detail
                        )
                    except:
                        pass
                raise HTTPException(status_code=422, detail=error_detail)

        # Save final result to database if user is authenticated
        if current_user and db_document:
            try:
                # Parse the JSON response
                parsed_data = json.loads(combined_response_text)

                # Save parsed data to database
                crud.update_document_parsed_data(
                    db=db,
                    document_id=document_id,
                    user_id=current_user.id,
                    parsed_data=parsed_data,
                    extraction_verified=safe_get(parsed_data, "extractionVerification", "extractionVerified", default=False),
                    verification_data=safe_get(parsed_data, "extractionVerification")
                )

                # Update status to completed
                crud.update_document_status(
                    db=db,
                    document_id=document_id,
                    user_id=current_user.id,
                    status="completed",
                    progress=100
                )

                # Extract and save transactions
                vendor_name = safe_get(parsed_data, "partyInformation", "vendor", "name")
                line_items = safe_get(parsed_data, "lineItems", default=[])
                if not isinstance(line_items, list):
                    line_items = []

                # If we have line items, save them as transactions
                if line_items:
                    for idx, item in enumerate(line_items):
                        try:
                            transaction_id = f"{document_id}-{idx}"
                            item_desc = item.get("description") if isinstance(item, dict) else None
                            item_price = item.get("totalPrice", 0) if isinstance(item, dict) else 0
                            crud.create_transaction(
                                db=db,
                                user_id=current_user.id,
                                document_id=db_document.id,
                                transaction_id=transaction_id,
                                vendor_name=vendor_name or item_desc,
                                amount=float(item_price or 0),
                                transaction_date=safe_get(parsed_data, "documentMetadata", "documentDate"),
                                description=item_desc,
                                transaction_type=safe_get(parsed_data, "documentMetadata", "documentType"),
                                line_items=[item] if isinstance(item, dict) else []
                            )
                        except Exception as e:
                            print(f"Warning: Failed to save transaction {idx}: {e}")

                # Log completion
                crud.log_activity(
                    db=db,
                    user_id=current_user.id,
                    action="document_processed",
                    entity_type="document",
                    entity_id=db_document.id,
                    details={"document_id": document_id, "transactions_count": len(line_items)}
                )
            except Exception as e:
                print(f"Warning: Failed to save processed data to database: {e}")
                # Update status to error
                if db_document:
                    try:
                        crud.update_document_status(
                            db=db,
                            document_id=document_id,
                            user_id=current_user.id,
                            status="error",
                            error_message=str(e)
                        )
                    except:
                        pass

        # Final validation - ensure we have valid response data
        if not combined_response_text or combined_response_text.strip() == "" or combined_response_text.strip() == "null":
            error_detail = "Document processing completed but no data was extracted. The document may be empty or unreadable."
            if current_user and db_document:
                try:
                    crud.update_document_status(
                        db=db,
                        document_id=document_id,
                        user_id=current_user.id,
                        status="error",
                        error_message=error_detail
                    )
                except:
                    pass
            raise HTTPException(status_code=422, detail=error_detail)

        # Validate that response is valid JSON before returning
        try:
            json.loads(combined_response_text)
        except json.JSONDecodeError:
            error_detail = "Document processing completed but response is not valid JSON."
            if current_user and db_document:
                try:
                    crud.update_document_status(
                        db=db,
                        document_id=document_id,
                        user_id=current_user.id,
                        status="error",
                        error_message=error_detail
                    )
                except:
                    pass
            raise HTTPException(status_code=422, detail=error_detail)

        # Return the merged Gemini response with document ID
        return {
            "response": combined_response_text.strip(),
            "document_id": document_id if current_user else None
        }
    except HTTPException:
        # Re-raise HTTP exceptions as-is
        raise
    except Exception as e:
        # Update document status to error if user is authenticated
        if current_user and db_document:
            try:
                crud.update_document_status(
                    db=db,
                    document_id=document_id,
                    user_id=current_user.id,
                    status="error",
                    error_message=str(e)
                )
            except:
                pass

        return {"error": "Request failed", "detail": str(e)}

# Define request model for vendor research
class VendorResearchRequest(BaseModel):
    vendor_name: str = Field(..., min_length=1, max_length=500, description="Vendor name to research")

    @field_validator('vendor_name')
    @classmethod
    def validate_vendor_name(cls, v):
        if not v or not v.strip():
            raise ValueError("Vendor name cannot be empty or whitespace")
        return v.strip()

@app.post("/research-vendor")
async def research_vendor(
    request: VendorResearchRequest,
    current_user: models.User = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """
    Research a vendor using Google Search.

    Results are cached in database if user is authenticated to avoid redundant API calls.
    """
    vendor_name = request.vendor_name

    if not vendor_name:
        return {"error": "No vendor name provided"}

    # Check cache if user is authenticated
    if current_user:
        try:
            cached_research = crud.get_or_create_vendor_research(
                db=db,
                user_id=current_user.id,
                vendor_name=vendor_name,
                fetch_if_missing=False  # Just check cache first
            )

            if cached_research and cached_research.research_data:
                # Return cached result
                print(f"Returning cached vendor research for: {vendor_name}")
                return {"response": cached_research.research_data.get("response", "")}
        except Exception as e:
            print(f"Warning: Failed to check vendor research cache: {e}")

    try:
        # Create a specific prompt asking for the single most likely entity and detailed info about it
        prompt = f"""
        Research the vendor "{vendor_name}" and identify the SINGLE most likely business or entity that this name refers to. 
        
        IMPORTANT: DO NOT list multiple possible interpretations or multiple businesses.
        Determine the most probable, prominent, or common business that matches this name and ONLY provide information about that ONE business.
        
        For this single most likely business, provide as much detail as possible about:
        - What this company does (main business focus)
        - Their products or services in detail
        - Company size, scale of operations, and market position
        - Company history and important milestones
        - Locations, reach, or distribution
        - Any other relevant details that would be helpful to know
        
        Again, I want information about the single most likely match only, not a list of possibilities.
        """
        
        # Use Google Search as a tool for grounding
        google_search_tool = types.Tool(
            google_search=types.GoogleSearch()
        )

        # Send the request to Gemini API with search enabled (with retry for rate limits)
        async def make_api_call():
            return await asyncio.to_thread(
                client.models.generate_content,
                model="gemini-2.0-flash",
                contents=prompt,
                config=types.GenerateContentConfig(
                    tools=[google_search_tool],
                    response_modalities=["TEXT"],
                    temperature=0.2,  # Lower temperature to make response more focused
                )
            )

        response = await retry_with_backoff(make_api_call, max_retries=3, initial_delay=2)

        # Save result to database if user is authenticated
        if current_user:
            try:
                crud.save_vendor_research(
                    db=db,
                    user_id=current_user.id,
                    vendor_name=vendor_name,
                    research_data={"response": response.text},
                    company_name=vendor_name,  # Could parse from response
                    description=response.text[:500] if len(response.text) > 500 else response.text
                )

                # Log activity
                crud.log_activity(
                    db=db,
                    user_id=current_user.id,
                    action="vendor_researched",
                    entity_type="vendor_research",
                    details={"vendor_name": vendor_name}
                )
            except Exception as e:
                print(f"Warning: Failed to save vendor research to database: {e}")

        # Return the response as-is
        return {"response": response.text}

    except Exception as e:
        print(f"Error researching vendor: {str(e)}")
        # Return user-friendly error message
        friendly_error = get_user_friendly_error(e)
        return {"error": friendly_error}

# Enhanced vendor research for ambiguous cases
class EnhancedResearchRequest(BaseModel):
    vendor_name: str = Field(..., min_length=1, max_length=500, description="Vendor name to research")
    transaction_context: dict = Field(..., description="Additional context about the transaction")
    confidence_threshold: int = Field(70, ge=0, le=100, description="Minimum confidence to skip research (0-100)")

    @field_validator('vendor_name')
    @classmethod
    def validate_vendor_name(cls, v):
        if not v or not v.strip():
            raise ValueError("Vendor name cannot be empty or whitespace")
        return v.strip()


@app.post("/research-vendor-enhanced")
async def research_vendor_enhanced(
    body: EnhancedResearchRequest,
    current_user: models.User = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """
    Enhanced vendor research for ambiguous or low-confidence transactions.

    Uses multiple analysis approaches:
    1. Industry classification
    2. Business type identification
    3. Typical transaction categorization
    4. Confidence scoring
    5. Red flag detection

    Results cached in database if user is authenticated.
    """
    vendor_name = body.vendor_name

    if not vendor_name:
        return {"error": "No vendor name provided"}

    # Check cache if user is authenticated
    if current_user:
        try:
            cached_research = crud.get_or_create_vendor_research(
                db=db,
                user_id=current_user.id,
                vendor_name=vendor_name,
                fetch_if_missing=False
            )

            if cached_research and cached_research.research_data:
                # Check if we have enhanced research data
                if cached_research.research_data.get("enhanced"):
                    print(f"Returning cached enhanced research for: {vendor_name}")
                    return cached_research.research_data
        except Exception as e:
            print(f"Warning: Failed to check research cache: {e}")

    try:
        # Enhanced research prompt with multiple analysis dimensions
        prompt = f"""
        Perform a comprehensive analysis of the vendor "{vendor_name}" for financial categorization purposes.

        Provide analysis in the following JSON structure:
        {{
            "vendorIdentification": {{
                "primaryName": "Most likely official business name",
                "aliases": ["Alternative names or abbreviations"],
                "confidence": "Score 0-100 on vendor identification certainty"
            }},
            "businessProfile": {{
                "industry": "Primary industry classification",
                "businessType": "Type of business (e.g., retailer, service provider, manufacturer)",
                "products": ["Main products or services offered"],
                "scale": "Business scale (local, regional, national, international)",
                "publiclyTraded": "Boolean - is this a public company"
            }},
            "categorizationGuidance": {{
                "typicalCategories": ["Most common expense categories for transactions with this vendor"],
                "reasoning": "Why these categories apply",
                "confidence": "Score 0-100 on categorization guidance certainty"
            }},
            "transactionContext": {{
                "commonTransactionTypes": ["Types of transactions typically made with this vendor"],
                "typicalAmountRanges": "Typical transaction amount ranges",
                "frequency": "How often businesses typically transact with this vendor"
            }},
            "ambiguityFactors": {{
                "nameClarity": "Score 0-100: How clear/unambiguous is the vendor name (100 = very clear)",
                "multipleMatches": "Boolean - could this name match multiple different businesses",
                "genericName": "Boolean - is this a generic/common business name",
                "industryAmbiguity": "Boolean - could this vendor operate in multiple industries"
            }},
            "redFlags": {{
                "hasFlags": "Boolean - any concerns detected",
                "flags": ["List of any concerns or unusual factors"],
                "severity": "low/medium/high"
            }},
            "overallConfidence": "Score 0-100: Overall confidence in this research",
            "recommendedAction": "What should be done next: 'accept' (high confidence), 'review' (medium confidence), 'manual_review' (low confidence)",
            "summary": "2-3 sentence summary of the vendor and key considerations for categorization"
        }}

        Transaction Context (if available):
        {json.dumps(body.transaction_context, indent=2) if body.transaction_context else "No additional context"}

        IMPORTANT: Be conservative with confidence scores. If there's any ambiguity, indicate it clearly.
        """

        # Use Google Search as a tool for grounding
        google_search_tool = types.Tool(
            google_search=types.GoogleSearch()
        )

        # Send the request to Gemini API with search enabled (with retry for rate limits)
        async def make_enhanced_api_call():
            return await asyncio.to_thread(
                client.models.generate_content,
                model="gemini-2.0-flash",
                contents=prompt,
                config=types.GenerateContentConfig(
                    tools=[google_search_tool],
                    response_modalities=["TEXT"],
                    response_mime_type="application/json",
                    temperature=0.2,
                )
            )

        response = await retry_with_backoff(make_enhanced_api_call, max_retries=3, initial_delay=2)

        # Parse the response
        try:
            research_data = json.loads(response.text)
            research_data["enhanced"] = True  # Mark as enhanced research
            research_data["vendor_name"] = vendor_name
            research_data["timestamp"] = datetime.utcnow().isoformat()

            # Save to database if user is authenticated
            if current_user:
                try:
                    crud.save_vendor_research(
                        db=db,
                        user_id=current_user.id,
                        vendor_name=vendor_name,
                        research_data=research_data,
                        company_name=safe_get(research_data, "vendorIdentification", "primaryName", default=vendor_name),
                        description=research_data.get("summary", "") if isinstance(research_data, dict) else "",
                        confidence_score=research_data.get("overallConfidence", 0) if isinstance(research_data, dict) else 0
                    )

                    # Log activity
                    crud.log_activity(
                        db=db,
                        user_id=current_user.id,
                        action="enhanced_vendor_research",
                        entity_type="vendor_research",
                        details={
                            "vendor_name": vendor_name,
                            "confidence": research_data.get("overallConfidence"),
                            "recommended_action": research_data.get("recommendedAction")
                        }
                    )
                except Exception as e:
                    print(f"Warning: Failed to save enhanced research: {e}")

            return research_data

        except json.JSONDecodeError:
            # If response isn't valid JSON, return error
            return {
                "error": "Failed to parse research response",
                "raw_response": response.text,
                "enhanced": True
            }

    except Exception as e:
        print(f"Error in enhanced vendor research: {str(e)}")
        # Return user-friendly error message
        friendly_error = get_user_friendly_error(e)
        return {"error": friendly_error}


# Define request model for financial categorization
class FinancialCategorizationRequest(BaseModel):
    vendor_info: str = Field(..., min_length=1, max_length=500, description="Vendor information")
    document_data: dict = Field(..., description="Transaction document data")
    transaction_purpose: str = Field("", max_length=1000, description="Transaction purpose")

    @field_validator('vendor_info')
    @classmethod
    def validate_vendor_info(cls, v):
        if not v or not v.strip():
            raise ValueError("Vendor info cannot be empty or whitespace")
        return v.strip()

@app.post("/categorize-transaction")
async def categorize_transaction(request: FinancialCategorizationRequest):
    vendor_info = request.vendor_info
    document_data = request.document_data
    transaction_purpose = request.transaction_purpose
    
    if not vendor_info or not document_data:
        return {"error": "Missing required information"}
    
    try:
        # Create a prompt that includes the categorization options and asks Gemini to categorize the transaction
        prompt = f"""
        Based on the information below, please categorize this transaction according to accounting principles.

        CRITICAL INSTRUCTION: You must categorize from the perspective of the INVOICE RECIPIENT (the customer being billed), NOT from the vendor's perspective.

        For example:
        - If this is an invoice FROM a vendor TO our business, it should typically be categorized as an Expense, Asset, or Liability
        - If this is a receipt we issued TO a customer FROM our business, it would be categorized as Revenue

        This categorization is for the accounting records of the business RECEIVING the invoice/document.

        Return a JSON object with the following structure:
        {{
            "companyName": "The name of the company that issued the invoice (the vendor)",
            "description": "A detailed description of what this business does",
            "category": "The most appropriate accounting category from the list below",
            "subcategory": "The most appropriate subcategory",
            "ledgerType": "The ledger entry type",
            "confidence": "A confidence score from 0-100 indicating how certain you are about this categorization. Consider factors like: clarity of vendor name, specificity of transaction details, alignment with typical business patterns, and ambiguity in the data.",
            "confidenceFactors": {{
                "vendorClarity": "Score 0-100: How clear and unambiguous is the vendor name",
                "dataCompleteness": "Score 0-100: How complete is the transaction data",
                "categoryFit": "Score 0-100: How well does the transaction fit the chosen category",
                "ambiguityLevel": "Score 0-100: Overall ambiguity (100 = very clear, 0 = very ambiguous)"
            }},
            "explanation": "A detailed explanation of why this categorization was chosen, including the factors considered and accounting principles applied",
            "needsResearch": "Boolean - true if additional vendor research would significantly improve categorization confidence"
        }}
        
        Here are the available categories, subcategories, and ledger types:
        
        Parent Category | Subcategory | Ledger Entry Type
        ---------------|-------------|------------------
        Revenue | Product Sales | Revenue
        Revenue | Service Revenue | Revenue
        Revenue | Rental Revenue | Revenue
        Revenue | Commission Revenue | Revenue
        Revenue | Subscription Revenue | Revenue
        Revenue | Other Income | Revenue
        Cost of Goods Sold (COGS) | Raw Materials | Expense (COGS)
        Cost of Goods Sold (COGS) | Direct Labor | Expense (COGS)
        Cost of Goods Sold (COGS) | Manufacturing Overhead | Expense (COGS)
        Cost of Goods Sold (COGS) | Freight and Delivery | Expense (COGS)
        Operating Expenses | Salaries and Wages | Expense (Operating)
        Operating Expenses | Rent | Expense (Operating)
        Operating Expenses | Utilities | Expense (Operating)
        Operating Expenses | Office Supplies | Expense (Operating)
        Operating Expenses | Business Software / IT Expenses | Expense (Operating)
        Operating Expenses | HR Expenses | Expense (Operating)
        Operating Expenses | Marketing and Advertising | Expense (Operating)
        Operating Expenses | Travel and Entertainment | Expense (Operating)
        Operating Expenses | Insurance | Expense (Operating)
        Operating Expenses | Repairs and Maintenance | Expense (Operating)
        Operating Expenses | Depreciation | Expense (Operating)
        Administrative Expenses | Professional Fees | Expense (Administrative)
        Administrative Expenses | Office Expenses | Expense (Administrative)
        Administrative Expenses | Postage and Shipping | Expense (Administrative)
        Administrative Expenses | Communication Expense | Expense (Administrative)
        Administrative Expenses | Bank Fees and Charges | Expense (Administrative)
        Financial Expenses | Interest Expense | Expense (Financial)
        Financial Expenses | Loan Fees | Expense (Financial)
        Financial Expenses | Credit Card Fees | Expense (Financial)
        Other Expenses | Miscellaneous | Expense (Other)
        Other Expenses | Donations/Charitable Contributions | Expense (Other)
        Other Expenses | Loss on Disposal of Assets | Expense (Other)
        Assets – Current | Cash and Cash Equivalents | Asset (Current)
        Assets – Current | Accounts Receivable | Asset (Current)
        Assets – Current | Inventory | Asset (Current)
        Assets – Current | Prepaid Expenses | Asset (Current)
        Assets – Current | Short-term Investments | Asset (Current)
        Assets – Fixed / Long-term | Property, Plant, and Equipment | Asset (Fixed)
        Assets – Fixed / Long-term | Furniture and Fixtures | Asset (Fixed)
        Assets – Fixed / Long-term | Vehicles | Asset (Fixed)
        Assets – Fixed / Long-term | Machinery and Equipment | Asset (Fixed)
        Assets – Fixed / Long-term | Computer Equipment | Asset (Fixed)
        Assets – Intangible | Patents | Asset (Intangible)
        Assets – Intangible | Trademarks | Asset (Intangible)
        Assets – Intangible | Copyrights | Asset (Intangible)
        Assets – Intangible | Goodwill | Asset (Intangible)
        Assets – Intangible | Capitalized Software | Asset (Intangible)
        Liabilities – Current | Accounts Payable | Liability (Current)
        Liabilities – Current | Short-term Loans | Liability (Current)
        Liabilities – Current | Accrued Liabilities | Liability (Current)
        Liabilities – Current | Current Portion of Long-term Debt | Liability (Current)
        Liabilities – Long-term | Long-term Loans | Liability (Long-term)
        Liabilities – Long-term | Bonds Payable | Liability (Long-term)
        Liabilities – Long-term | Deferred Tax Liabilities | Liability (Long-term)
        Equity | Common Stock | Equity
        Equity | Retained Earnings | Equity
        Equity | Additional Paid-in Capital | Equity
        Equity | Dividends/Distributions | Equity
        Adjusting / Journal Entries | Accruals/Deferrals/Depreciation Adjustments | Adjustment
        
        Vendor Information (the seller/company that sent the invoice):
        {vendor_info}
        
        Document Data:
        {json.dumps(document_data, indent=2)}
        
        Transaction Purpose (what the invoice is for):
        {transaction_purpose}
        
        REMEMBER: Categorize from the perspective of the business RECEIVING this invoice - the customer being billed, NOT from the perspective of the vendor who issued it.
        
        In the explanation field, be thorough about why this specific category, subcategory, and ledger type was chosen. 
        Consider the nature of the transaction, the items or services involved, accounting best practices, and how this 
        classification aligns with standard chart of accounts structures.
        """
        
        # Send the request to Gemini API
        response = await asyncio.to_thread(
            client.models.generate_content,
            model="gemini-2.0-flash",
            contents=prompt,
            config={
                "max_output_tokens": 4000,
                "response_mime_type": "application/json"
            }
        )
        
        # Return the response
        try:
            # Try to parse the response as JSON
            categorization_json = json.loads(response.text)
            return {"response": categorization_json}
        except json.JSONDecodeError:
            # If it's not valid JSON, return the raw text
            return {"response": response.text}
            
    except Exception as e:
        print(f"Error categorizing transaction: {str(e)}")
        return {"error": f"Error categorizing transaction: {str(e)}"}

# Smart categorization with automatic ambiguity resolution
class SmartCategorizationRequest(BaseModel):
    vendor_name: str = Field(..., min_length=1, max_length=500, description="Vendor name")
    document_data: dict = Field(..., description="Transaction document data")
    transaction_purpose: str = Field("", max_length=1000, description="Transaction purpose")
    confidence_threshold: int = Field(70, ge=0, le=100, description="Confidence threshold (0-100)")
    auto_research: bool = Field(True, description="Auto-trigger research on low confidence")

    @field_validator('vendor_name')
    @classmethod
    def validate_vendor_name(cls, v):
        if not v or not v.strip():
            raise ValueError("Vendor name cannot be empty or whitespace")
        return v.strip()


@app.post("/categorize-transaction-smart")
async def categorize_transaction_smart(
    body: SmartCategorizationRequest,
    current_user: models.User = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """
    Smart categorization with automatic ambiguity resolution.

    Workflow:
    1. Perform initial categorization
    2. Check confidence score
    3. If confidence < threshold AND auto_research=true: trigger enhanced vendor research
    4. Re-categorize with enhanced context
    5. Flag for manual review if still low confidence
    6. Return comprehensive results with all confidence metrics

    This endpoint provides the most intelligent categorization with built-in ambiguity handling.
    """
    vendor_name = body.vendor_name
    document_data = body.document_data
    transaction_purpose = body.transaction_purpose

    if not vendor_name or not document_data:
        return {"error": "Missing required information"}

    result = {
        "vendor_name": vendor_name,
        "workflow": [],
        "final_categorization": None,
        "confidence_metrics": {},
        "research_performed": False,
        "needs_manual_review": False,
        "timestamp": datetime.utcnow().isoformat()
    }

    try:
        # Step 1: Initial categorization
        result["workflow"].append({
            "step": 1,
            "action": "initial_categorization",
            "status": "in_progress"
        })

        # Get initial categorization using Gemini
        initial_categorization = await _get_gemini_categorization(
            vendor_name,
            document_data,
            transaction_purpose
        )

        result["initial_categorization"] = initial_categorization
        result["workflow"][-1]["status"] = "completed"

        # Extract confidence score (ensure it's a number)
        raw_confidence = initial_categorization.get("confidence", 0)
        try:
            if isinstance(raw_confidence, str):
                confidence = float(''.join(c for c in raw_confidence if c.isdigit() or c == '.') or '0')
            else:
                confidence = float(raw_confidence) if raw_confidence else 0
        except (ValueError, TypeError):
            confidence = 0
        result["confidence_metrics"]["initial_confidence"] = confidence

        # Step 2: Check if enhanced research is needed
        needs_research = (
            confidence < body.confidence_threshold or
            initial_categorization.get("needsResearch", False)
        )

        result["workflow"].append({
            "step": 2,
            "action": "confidence_check",
            "confidence": confidence,
            "threshold": body.confidence_threshold,
            "needs_research": needs_research,
            "status": "completed"
        })

        # Step 3: Enhanced research if needed
        if needs_research and body.auto_research:
            result["workflow"].append({
                "step": 3,
                "action": "enhanced_research",
                "status": "in_progress"
            })

            try:
                # Perform enhanced vendor research
                enhanced_research = await research_vendor_enhanced(
                    EnhancedResearchRequest(
                        vendor_name=vendor_name,
                        transaction_context=document_data,
                        confidence_threshold=body.confidence_threshold
                    ),
                    current_user=current_user,
                    db=db
                )

                result["enhanced_research"] = enhanced_research
                result["research_performed"] = True
                result["workflow"][-1]["status"] = "completed"
                result["confidence_metrics"]["research_confidence"] = enhanced_research.get("overallConfidence", 0)

                # Step 4: Re-categorize with enhanced context
                result["workflow"].append({
                    "step": 4,
                    "action": "re_categorization_with_research",
                    "status": "in_progress"
                })

                # Create enhanced vendor info string
                typical_cats = safe_get(enhanced_research, 'categorizationGuidance', 'typicalCategories', default=[])
                typical_cats_str = ', '.join(typical_cats) if isinstance(typical_cats, list) else str(typical_cats)
                enhanced_vendor_info = f"""
                Vendor: {vendor_name}

                Research Findings:
                - Official Name: {safe_get(enhanced_research, 'vendorIdentification', 'primaryName', default=vendor_name)}
                - Industry: {safe_get(enhanced_research, 'businessProfile', 'industry', default='Unknown')}
                - Business Type: {safe_get(enhanced_research, 'businessProfile', 'businessType', default='Unknown')}
                - Summary: {enhanced_research.get('summary', 'No summary available') if isinstance(enhanced_research, dict) else 'No summary available'}
                - Typical Categories: {typical_cats_str}
                - Research Confidence: {enhanced_research.get('overallConfidence', 0) if isinstance(enhanced_research, dict) else 0}%
                """

                # Re-categorize with enhanced context
                final_categorization = await _get_gemini_categorization(
                    enhanced_vendor_info,
                    document_data,
                    transaction_purpose
                )

                result["final_categorization"] = final_categorization
                result["workflow"][-1]["status"] = "completed"
                # Ensure confidence is a number
                raw_final_conf = final_categorization.get("confidence", 0)
                try:
                    if isinstance(raw_final_conf, str):
                        final_conf_value = float(''.join(c for c in raw_final_conf if c.isdigit() or c == '.') or '0')
                    else:
                        final_conf_value = float(raw_final_conf) if raw_final_conf else 0
                except (ValueError, TypeError):
                    final_conf_value = 0
                result["confidence_metrics"]["final_confidence"] = final_conf_value

            except Exception as e:
                result["workflow"][-1]["status"] = "error"
                result["workflow"][-1]["error"] = str(e)
                # Fall back to initial categorization
                result["final_categorization"] = initial_categorization
                result["confidence_metrics"]["final_confidence"] = confidence

        else:
            # Use initial categorization as final
            result["final_categorization"] = initial_categorization
            result["confidence_metrics"]["final_confidence"] = confidence

        # Step 5: Determine if manual review is needed
        # Ensure final_confidence is a number for comparison
        raw_final = result["confidence_metrics"].get("final_confidence", 0)
        try:
            if isinstance(raw_final, str):
                final_confidence = float(''.join(c for c in raw_final if c.isdigit() or c == '.') or '0')
            else:
                final_confidence = float(raw_final) if raw_final else 0
        except (ValueError, TypeError):
            final_confidence = 0

        # Flag for manual review if:
        # - Final confidence is still below threshold, OR
        # - Research recommended manual review, OR
        # - Red flags were detected
        recommended_action = safe_get(result, "enhanced_research", "recommendedAction", default="")
        red_flag_severity = safe_get(result, "enhanced_research", "redFlags", "severity", default="")
        needs_manual_review = (
            final_confidence < body.confidence_threshold or
            (recommended_action == "manual_review") or
            (red_flag_severity in ["medium", "high"])
        )

        result["needs_manual_review"] = needs_manual_review

        result["workflow"].append({
            "step": 5,
            "action": "review_determination",
            "needs_manual_review": needs_manual_review,
            "final_confidence": final_confidence,
            "status": "completed"
        })

        # Save to database if user is authenticated
        if current_user:
            try:
                # Find the transaction in database
                transaction_id = document_data.get("id", "")
                if transaction_id:
                    db_transaction = crud.get_transaction_by_id(db, str(transaction_id), current_user.id)

                    if db_transaction:
                        # Save categorization with needs_review flag
                        categorization_data = {
                            "category": result["final_categorization"].get("category"),
                            "subcategory": result["final_categorization"].get("subcategory"),
                            "ledger_type": result["final_categorization"].get("ledgerType"),
                            "method": "smart_ai",
                            "confidence_score": final_confidence,
                            "explanation": result["final_categorization"].get("explanation"),
                            "transaction_purpose": transaction_purpose,
                            "full_result": result
                        }
                        db_cat = crud.create_categorization(
                            db=db,
                            user_id=current_user.id,
                            categorization_data=categorization_data,
                            transaction_id=db_transaction.id
                        )
                        # Auto-approve if high confidence
                        if not needs_manual_review:
                            crud.update_categorization_approval(
                                db, db_cat.id, current_user.id, approved=True
                            )

                        # Update transaction to flag for review if needed
                        if needs_manual_review:
                            db.query(models.Transaction).filter(
                                models.Transaction.id == db_transaction.id
                            ).update({"notes": f"NEEDS REVIEW - Confidence: {final_confidence}%"})
                            db.commit()

                        # Log activity
                        crud.log_activity(
                            db=db,
                            user_id=current_user.id,
                            action="smart_categorization",
                            entity_type="categorization",
                            details={
                                "vendor_name": vendor_name,
                                "final_confidence": final_confidence,
                                "research_performed": result["research_performed"],
                                "needs_manual_review": needs_manual_review
                            }
                        )
            except Exception as e:
                print(f"Warning: Failed to save smart categorization: {e}")

        return {
            "success": True,
            **result
        }

    except Exception as e:
        print(f"Error in smart categorization: {str(e)}")
        return {
            "success": False,
            "error": f"Error in smart categorization: {str(e)}"
        }


# Define request model for hybrid categorization
class HybridCategorizationRequest(BaseModel):
    vendor_info: str
    document_data: dict
    transaction_purpose: str = ""

@app.post("/categorize-transaction-hybrid")
async def categorize_transaction_hybrid(request: HybridCategorizationRequest):
    """
    Hybrid categorization endpoint that provides BOTH ML prediction and Gemini AI categorization.

    This allows users to compare both approaches and choose the best one, while also
    enabling the system to learn from their choices.
    """
    vendor_info = request.vendor_info
    document_data = request.document_data
    transaction_purpose = request.transaction_purpose

    if not vendor_info or not document_data:
        return {"error": "Missing required information"}

    # Ensure document_data is a dict, not a list
    if isinstance(document_data, list):
        if len(document_data) > 0 and isinstance(document_data[0], dict):
            document_data = document_data[0]  # Take first item
        else:
            return {"error": "Invalid document data format - expected object, got list"}

    if not isinstance(document_data, dict):
        return {"error": "Invalid document data format - expected object"}

    try:
        # Get ML engine
        engine = get_ml_categorization_engine()

        # Run ML prediction and Gemini categorization in parallel for speed
        ml_prediction_task = engine.predict_category(
            document_data,
            transaction_purpose
        )

        # Gemini categorization (existing logic)
        gemini_categorization_task = _get_gemini_categorization(
            vendor_info,
            document_data,
            transaction_purpose
        )

        # Wait for both to complete
        ml_prediction, gemini_categorization = await asyncio.gather(
            ml_prediction_task,
            gemini_categorization_task
        )

        # Return both predictions
        return {
            "mlPrediction": ml_prediction,
            "geminiCategorization": gemini_categorization,
            "hybridApproach": True,
            "timestamp": asyncio.get_event_loop().time()
        }

    except ValueError as ve:
        # ML engine not initialized (likely missing Pinecone API key)
        print(f"ML engine not available: {str(ve)}")

        # Fallback to Gemini only
        gemini_categorization = await _get_gemini_categorization(
            vendor_info,
            document_data,
            transaction_purpose
        )

        return {
            "mlPrediction": {
                "hasPrediction": False,
                "confidence": 0.0,
                "reason": "ML engine not configured. Please add PINECONE_API_KEY to .env file."
            },
            "geminiCategorization": gemini_categorization,
            "hybridApproach": False,
            "fallbackMode": "gemini_only"
        }

    except Exception as e:
        print(f"Error in hybrid categorization: {str(e)}")
        return {"error": f"Error in hybrid categorization: {str(e)}"}

async def _get_gemini_categorization(vendor_info: str, document_data: dict, transaction_purpose: str) -> dict:
    """
    Helper function to get Gemini AI categorization (extracted from existing endpoint).
    """
    # Create a prompt that includes the categorization options and asks Gemini to categorize the transaction
    prompt = f"""
    Based on the information below, please categorize this transaction according to accounting principles.

    CRITICAL INSTRUCTION: You must categorize from the perspective of the INVOICE RECIPIENT (the customer being billed), NOT from the vendor's perspective.

    For example:
    - If this is an invoice FROM a vendor TO our business, it should typically be categorized as an Expense, Asset, or Liability
    - If this is a receipt we issued TO a customer FROM our business, it would be categorized as Revenue

    This categorization is for the accounting records of the business RECEIVING the invoice/document.

    Return a JSON object with the following structure:
    {{
        "companyName": "The name of the company that issued the invoice (the vendor)",
        "description": "A detailed description of what this business does",
        "category": "The most appropriate accounting category from the list below",
        "subcategory": "The most appropriate subcategory",
        "ledgerType": "The ledger entry type",
        "confidence": "A confidence score from 0-100 indicating how certain you are about this categorization. Consider factors like: clarity of vendor name, specificity of transaction details, alignment with typical business patterns, and ambiguity in the data.",
        "confidenceFactors": {{
            "vendorClarity": "Score 0-100: How clear and unambiguous is the vendor name",
            "dataCompleteness": "Score 0-100: How complete is the transaction data",
            "categoryFit": "Score 0-100: How well does the transaction fit the chosen category",
            "ambiguityLevel": "Score 0-100: Overall ambiguity (100 = very clear, 0 = very ambiguous)"
        }},
        "explanation": "A detailed explanation of why this categorization was chosen, including the factors considered and accounting principles applied",
        "needsResearch": "Boolean - true if additional vendor research would significantly improve categorization confidence"
    }}

    Here are the available categories, subcategories, and ledger types:

    Parent Category | Subcategory | Ledger Entry Type
    ---------------|-------------|------------------
    Revenue | Product Sales | Revenue
    Revenue | Service Revenue | Revenue
    Revenue | Rental Revenue | Revenue
    Revenue | Commission Revenue | Revenue
    Revenue | Subscription Revenue | Revenue
    Revenue | Other Income | Revenue
    Cost of Goods Sold (COGS) | Raw Materials | Expense (COGS)
    Cost of Goods Sold (COGS) | Direct Labor | Expense (COGS)
    Cost of Goods Sold (COGS) | Manufacturing Overhead | Expense (COGS)
    Cost of Goods Sold (COGS) | Freight and Delivery | Expense (COGS)
    Operating Expenses | Salaries and Wages | Expense (Operating)
    Operating Expenses | Rent | Expense (Operating)
    Operating Expenses | Utilities | Expense (Operating)
    Operating Expenses | Office Supplies | Expense (Operating)
    Operating Expenses | Business Software / IT Expenses | Expense (Operating)
    Operating Expenses | HR Expenses | Expense (Operating)
    Operating Expenses | Marketing and Advertising | Expense (Operating)
    Operating Expenses | Travel and Entertainment | Expense (Operating)
    Operating Expenses | Insurance | Expense (Operating)
    Operating Expenses | Repairs and Maintenance | Expense (Operating)
    Operating Expenses | Depreciation | Expense (Operating)
    Administrative Expenses | Professional Fees | Expense (Administrative)
    Administrative Expenses | Office Expenses | Expense (Administrative)
    Administrative Expenses | Postage and Shipping | Expense (Administrative)
    Administrative Expenses | Communication Expense | Expense (Administrative)
    Administrative Expenses | Bank Fees and Charges | Expense (Administrative)
    Financial Expenses | Interest Expense | Expense (Financial)
    Financial Expenses | Loan Fees | Expense (Financial)
    Financial Expenses | Credit Card Fees | Expense (Financial)
    Other Expenses | Miscellaneous | Expense (Other)
    Other Expenses | Donations/Charitable Contributions | Expense (Other)
    Other Expenses | Loss on Disposal of Assets | Expense (Other)
    Assets – Current | Cash and Cash Equivalents | Asset (Current)
    Assets – Current | Accounts Receivable | Asset (Current)
    Assets – Current | Inventory | Asset (Current)
    Assets – Current | Prepaid Expenses | Asset (Current)
    Assets – Current | Short-term Investments | Asset (Current)
    Assets – Fixed / Long-term | Property, Plant, and Equipment | Asset (Fixed)
    Assets – Fixed / Long-term | Furniture and Fixtures | Asset (Fixed)
    Assets – Fixed / Long-term | Vehicles | Asset (Fixed)
    Assets – Fixed / Long-term | Machinery and Equipment | Asset (Fixed)
    Assets – Fixed / Long-term | Computer Equipment | Asset (Fixed)
    Assets – Intangible | Patents | Asset (Intangible)
    Assets – Intangible | Trademarks | Asset (Intangible)
    Assets – Intangible | Copyrights | Asset (Intangible)
    Assets – Intangible | Goodwill | Asset (Intangible)
    Assets – Intangible | Capitalized Software | Asset (Intangible)
    Liabilities – Current | Accounts Payable | Liability (Current)
    Liabilities – Current | Short-term Loans | Liability (Current)
    Liabilities – Current | Accrued Liabilities | Liability (Current)
    Liabilities – Current | Current Portion of Long-term Debt | Liability (Current)
    Liabilities – Long-term | Long-term Loans | Liability (Long-term)
    Liabilities – Long-term | Bonds Payable | Liability (Long-term)
    Liabilities – Long-term | Deferred Tax Liabilities | Liability (Long-term)
    Equity | Common Stock | Equity
    Equity | Retained Earnings | Equity
    Equity | Additional Paid-in Capital | Equity
    Equity | Dividends/Distributions | Equity
    Adjusting / Journal Entries | Accruals/Deferrals/Depreciation Adjustments | Adjustment

    Vendor Information (the seller/company that sent the invoice):
    {vendor_info}

    Document Data:
    {json.dumps(document_data, indent=2)}

    Transaction Purpose (what the invoice is for):
    {transaction_purpose}

    REMEMBER: Categorize from the perspective of the business RECEIVING this invoice - the customer being billed, NOT from the perspective of the vendor who issued it.

    In the explanation field, be thorough about why this specific category, subcategory, and ledger type was chosen.
    Consider the nature of the transaction, the items or services involved, accounting best practices, and how this
    classification aligns with standard chart of accounts structures.
    """

    # Send the request to Gemini API
    response = await asyncio.to_thread(
        client.models.generate_content,
        model="gemini-2.0-flash",
        contents=prompt,
        config={
            "max_output_tokens": 4000,
            "response_mime_type": "application/json"
        }
    )

    # Parse and return the response
    try:
        categorization_json = json.loads(response.text)

        # Handle case where Gemini returns a list instead of dict
        if isinstance(categorization_json, list):
            if len(categorization_json) > 0 and isinstance(categorization_json[0], dict):
                categorization_json = categorization_json[0]  # Take first item if it's a dict
            else:
                return {"error": "Unexpected response format from AI", "confidence": 0}

        # Ensure we have a dict
        if not isinstance(categorization_json, dict):
            return {"error": "Invalid response format from AI", "confidence": 0}

        # Ensure confidence is a valid number (Gemini sometimes returns it as string)
        if "confidence" in categorization_json:
            try:
                conf_value = categorization_json["confidence"]
                if isinstance(conf_value, str):
                    # Remove any non-numeric characters and convert
                    conf_value = ''.join(c for c in conf_value if c.isdigit() or c == '.')
                categorization_json["confidence"] = float(conf_value) if conf_value else 50
            except (ValueError, TypeError):
                categorization_json["confidence"] = 50  # Default to 50% if parsing fails
        else:
            categorization_json["confidence"] = 50  # Default confidence if missing

        return categorization_json
    except json.JSONDecodeError:
        return {"error": "Failed to parse Gemini response", "rawText": response.text, "confidence": 0}

# Define request model for storing categorization
class StoreCategorizationRequest(BaseModel):
    transaction_data: dict
    categorization: dict
    transaction_purpose: str = ""
    selected_method: str = "gemini"  # "ml", "gemini", or "manual"
    user_feedback: str = ""  # Optional feedback

@app.post("/store-categorization")
async def store_categorization(
    request: StoreCategorizationRequest,
    current_user: models.User = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """
    Store a categorization decision for machine learning.

    This endpoint is called after the user selects their preferred categorization
    (either ML, Gemini, or manually adjusted). The system learns from this decision.
    Saves to both Pinecone (for ML) and PostgreSQL (for history and audit).
    """
    try:
        # Get ML engine
        engine = get_ml_categorization_engine()

        # Prepare user feedback string
        feedback = f"{request.selected_method}"
        if request.user_feedback:
            feedback += f" | {request.user_feedback}"

        # Store in vector database (Pinecone)
        transaction_id = await engine.store_transaction(
            transaction_data=request.transaction_data,
            categorization=request.categorization,
            transaction_purpose=request.transaction_purpose,
            user_feedback=feedback
        )

        # Also save to PostgreSQL database if user is authenticated
        if current_user:
            try:
                # Find the transaction in our database (if it exists)
                db_transaction = crud.get_transaction_by_id(
                    db, str(request.transaction_data.get("id", "")), current_user.id
                )

                if db_transaction:
                    # Save categorization
                    categorization_data = {
                        "category": request.categorization.get("category"),
                        "subcategory": request.categorization.get("subcategory"),
                        "ledger_type": request.categorization.get("ledgerType"),
                        "method": request.selected_method,
                        "confidence_score": request.categorization.get("confidence", 0),
                        "explanation": request.categorization.get("explanation"),
                        "transaction_purpose": request.transaction_purpose,
                        "full_categorization": request.categorization
                    }
                    db_cat = crud.create_categorization(
                        db=db,
                        user_id=current_user.id,
                        categorization_data=categorization_data,
                        transaction_id=db_transaction.id
                    )
                    # Mark as approved since user selected it
                    crud.update_categorization_approval(
                        db, db_cat.id, current_user.id, approved=True
                    )

                    # Log activity
                    crud.log_activity(
                        db=db,
                        user_id=current_user.id,
                        action="categorization_saved",
                        entity_type="categorization",
                        details={
                            "method": request.selected_method,
                            "category": request.categorization.get("category")
                        }
                    )
            except Exception as e:
                print(f"Warning: Failed to save categorization to database: {e}")

        return {
            "success": True,
            "transactionId": transaction_id,
            "message": "Categorization stored successfully for future learning"
        }

    except ValueError as ve:
        return {
            "success": False,
            "error": "ML engine not configured",
            "detail": str(ve)
        }
    except Exception as e:
        print(f"Error storing categorization: {str(e)}")
        return {
            "success": False,
            "error": f"Error storing categorization: {str(e)}"
        }

@app.get("/ml-stats")
async def get_ml_stats():
    """
    Get statistics about the ML vector database.

    Returns information about how many transactions are stored and ready for learning.
    """
    try:
        # Get ML engine
        engine = get_ml_categorization_engine()

        # Get stats
        stats = await engine.get_database_stats()

        return {
            "success": True,
            "stats": stats
        }

    except ValueError as ve:
        return {
            "success": False,
            "error": "ML engine not configured",
            "detail": str(ve),
            "stats": {
                "totalTransactions": 0,
                "status": "not_configured"
            }
        }
    except Exception as e:
        print(f"Error getting ML stats: {str(e)}")
        return {
            "success": False,
            "error": f"Error getting ML stats: {str(e)}"
        }

# Define request model for submitting corrections
class SubmitCorrectionRequest(BaseModel):
    transaction_id: str = Field(..., min_length=1, max_length=100, description="Transaction ID")
    original_categorization: dict = Field(..., description="Original categorization data")
    corrected_categorization: dict = Field(..., description="Corrected categorization data")
    transaction_data: dict = Field(..., description="Transaction data")
    transaction_purpose: str = Field("", max_length=1000, description="Transaction purpose")
    correction_reason: str = Field("", max_length=2000, description="Reason for correction")

    @field_validator('transaction_id')
    @classmethod
    def validate_transaction_id(cls, v):
        if not v or not v.strip():
            raise ValueError("Transaction ID cannot be empty")
        return v.strip()

@app.post("/submit-correction")
async def submit_correction(
    request: SubmitCorrectionRequest,
    current_user: models.User = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """
    Submit a user correction to improve the ML model.

    This endpoint allows users to:
    1. Reject an incorrect ML prediction
    2. Manually correct a categorization
    3. Provide feedback on why the correction was needed

    The correction is stored with higher learning weight for future predictions.
    Saves to both Pinecone (for ML) and PostgreSQL (for history and analytics).
    """
    try:
        # Get ML engine
        engine = get_ml_categorization_engine()

        # Submit the correction to Pinecone
        result = await engine.submit_correction(
            transaction_id=request.transaction_id,
            original_categorization=request.original_categorization,
            corrected_categorization=request.corrected_categorization,
            transaction_data=request.transaction_data,
            transaction_purpose=request.transaction_purpose,
            correction_reason=request.correction_reason
        )

        # Also save to PostgreSQL database if user is authenticated
        if current_user:
            try:
                # Find the transaction in our database
                db_transaction = crud.get_transaction_by_id(
                    db, request.transaction_id, current_user.id
                )

                if db_transaction:
                    # Save user correction
                    crud.create_user_correction(
                        db=db,
                        user_id=current_user.id,
                        transaction_id=db_transaction.id,
                        original_category=request.original_categorization.get("category"),
                        original_subcategory=request.original_categorization.get("subcategory"),
                        original_ledger_type=request.original_categorization.get("ledgerType"),
                        original_method=request.original_categorization.get("method", "gemini"),
                        corrected_category=request.corrected_categorization.get("category"),
                        corrected_subcategory=request.corrected_categorization.get("subcategory"),
                        corrected_ledger_type=request.corrected_categorization.get("ledgerType"),
                        correction_reason=request.correction_reason,
                        correction_data={"original": request.original_categorization, "corrected": request.corrected_categorization}
                    )

                    # Log activity
                    crud.log_activity(
                        db=db,
                        user_id=current_user.id,
                        action="categorization_corrected",
                        entity_type="user_correction",
                        details={"transaction_id": request.transaction_id, "reason": request.correction_reason}
                    )
            except Exception as e:
                print(f"Warning: Failed to save correction to database: {e}")

        return result

    except ValueError as ve:
        return {
            "success": False,
            "error": "ML engine not configured",
            "detail": str(ve)
        }
    except Exception as e:
        print(f"Error submitting correction: {str(e)}")
        return {
            "success": False,
            "error": f"Error submitting correction: {str(e)}"
        }

@app.get("/correction-stats")
async def get_correction_stats():
    """
    Get statistics about corrections and learning progress.

    Returns insights about how many corrections have been made,
    learning status, and recommendations for improvement.
    """
    try:
        # Get ML engine
        engine = get_ml_categorization_engine()

        # Get correction stats
        stats = await engine.get_correction_stats()

        return {
            "success": True,
            "stats": stats
        }

    except ValueError as ve:
        return {
            "success": False,
            "error": "ML engine not configured",
            "detail": str(ve)
        }
    except Exception as e:
        print(f"Error getting correction stats: {str(e)}")
        return {
            "success": False,
            "error": f"Error getting correction stats: {str(e)}"
        }

@app.get("/insights/dashboard", tags=["Insights"])
async def get_insights_dashboard(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Get comprehensive insights and statistics for the categorization dashboard.

    Returns:
    - Overall categorization statistics
    - Category distribution
    - Confidence distribution
    - Approval rates
    - Learning progress metrics
    - Recent activity
    """
    try:
        from sqlalchemy import func, case
        from datetime import datetime, timedelta

        user_id = current_user.id

        # Get total categorizations for user
        total_categorizations = db.query(func.count(models.Categorization.id)).filter(
            models.Categorization.user_id == user_id
        ).scalar() or 0

        # Get approved vs pending
        approved_count = db.query(func.count(models.Categorization.id)).filter(
            models.Categorization.user_id == user_id,
            models.Categorization.user_approved == True
        ).scalar() or 0

        # Get user modified (corrections)
        corrections_count = db.query(func.count(models.Categorization.id)).filter(
            models.Categorization.user_id == user_id,
            models.Categorization.user_modified == True
        ).scalar() or 0

        # Get method distribution
        method_distribution = db.query(
            models.Categorization.method,
            func.count(models.Categorization.id).label('count')
        ).filter(
            models.Categorization.user_id == user_id
        ).group_by(models.Categorization.method).all()

        method_stats = {method: count for method, count in method_distribution}

        # Get category distribution (top 10)
        category_distribution = db.query(
            models.Categorization.category,
            func.count(models.Categorization.id).label('count'),
            func.sum(
                case(
                    (models.BankTransaction.amount != None, models.BankTransaction.amount),
                    else_=0
                )
            ).label('total_amount')
        ).outerjoin(
            models.BankTransaction,
            models.Categorization.bank_transaction_id == models.BankTransaction.id
        ).filter(
            models.Categorization.user_id == user_id
        ).group_by(
            models.Categorization.category
        ).order_by(
            func.count(models.Categorization.id).desc()
        ).limit(10).all()

        category_stats = [
            {
                "category": cat,
                "count": count,
                "total_amount": float(amount) if amount else 0
            }
            for cat, count, amount in category_distribution
        ]

        # Get confidence distribution
        confidence_ranges = db.query(
            case(
                (models.Categorization.confidence_score >= 90, 'high'),
                (models.Categorization.confidence_score >= 70, 'medium'),
                (models.Categorization.confidence_score >= 50, 'low'),
                else_='very_low'
            ).label('confidence_level'),
            func.count(models.Categorization.id).label('count')
        ).filter(
            models.Categorization.user_id == user_id,
            models.Categorization.confidence_score != None
        ).group_by('confidence_level').all()

        confidence_stats = {level: count for level, count in confidence_ranges}

        # Get average confidence
        avg_confidence = db.query(
            func.avg(models.Categorization.confidence_score)
        ).filter(
            models.Categorization.user_id == user_id,
            models.Categorization.confidence_score != None
        ).scalar()

        # Get bank statement stats
        total_statements = db.query(func.count(models.BankStatement.id)).filter(
            models.BankStatement.user_id == user_id
        ).scalar() or 0

        total_bank_transactions = db.query(func.count(models.BankTransaction.id)).filter(
            models.BankTransaction.user_id == user_id
        ).scalar() or 0

        # Get recent activity (last 7 days)
        seven_days_ago = datetime.utcnow() - timedelta(days=7)
        recent_categorizations = db.query(func.count(models.Categorization.id)).filter(
            models.Categorization.user_id == user_id,
            models.Categorization.created_at >= seven_days_ago
        ).scalar() or 0

        # Calculate approval rate
        approval_rate = (approved_count / total_categorizations * 100) if total_categorizations > 0 else 0

        # Calculate accuracy estimate (approved without modification)
        auto_approved = approved_count - corrections_count
        accuracy_estimate = (auto_approved / approved_count * 100) if approved_count > 0 else 0

        # Get ML stats if available
        ml_stats = None
        try:
            engine = get_ml_categorization_engine()
            ml_stats = await engine.get_database_stats()
        except:
            pass

        return {
            "success": True,
            "insights": {
                "overview": {
                    "total_categorizations": total_categorizations,
                    "approved_count": approved_count,
                    "pending_count": total_categorizations - approved_count,
                    "corrections_count": corrections_count,
                    "approval_rate": round(approval_rate, 1),
                    "accuracy_estimate": round(accuracy_estimate, 1),
                    "average_confidence": round(float(avg_confidence), 1) if avg_confidence else 0
                },
                "bank_statements": {
                    "total_statements": total_statements,
                    "total_transactions": total_bank_transactions
                },
                "method_distribution": method_stats,
                "category_distribution": category_stats,
                "confidence_distribution": confidence_stats,
                "recent_activity": {
                    "last_7_days": recent_categorizations
                },
                "ml_stats": ml_stats
            }
        }

    except Exception as e:
        print(f"Error getting insights dashboard: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error getting insights: {str(e)}"
        )


@app.get("/categories")
async def get_categories():
    """
    Get all available accounting categories for manual selection.

    Returns the complete list of categories, subcategories, and ledger types
    that can be used for categorization.
    """
    try:
        return {
            "success": True,
            "categories": get_all_categories(),
            "grouped": get_categories_by_parent()
        }
    except Exception as e:
        print(f"Error getting categories: {str(e)}")
        return {
            "success": False,
            "error": f"Error getting categories: {str(e)}"
        }

@app.get("/categories/{category}/subcategories")
async def get_subcategories(category: str):
    """
    Get subcategories for a specific parent category.

    Useful for cascading dropdowns where user first selects parent category,
    then subcategory.
    """
    try:
        subcategories = get_subcategories_for_category(category)
        return {
            "success": True,
            "category": category,
            "subcategories": subcategories
        }
    except Exception as e:
        print(f"Error getting subcategories: {str(e)}")
        return {
            "success": False,
            "error": f"Error getting subcategories: {str(e)}"
        }

# Bank Statement Reconciliation Endpoints

@app.get("/bank-statements", tags=["Bank Statements"])
async def list_bank_statements(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    List all bank statements for the current user.

    Returns statements ordered by most recent upload first.
    """
    statements = crud.get_bank_statements_by_user(db, current_user.id, skip, limit)

    return [
        {
            "id": stmt.id,
            "file_name": stmt.file_name,
            "file_type": stmt.file_type,
            "bank_name": stmt.bank_name,
            "account_number": stmt.account_number,
            "statement_date": str(stmt.statement_date) if stmt.statement_date else None,
            "period_start": str(stmt.period_start) if stmt.period_start else None,
            "period_end": str(stmt.period_end) if stmt.period_end else None,
            "transaction_count": stmt.transaction_count or 0,
            "uploaded_at": str(stmt.uploaded_at) if stmt.uploaded_at else None
        }
        for stmt in statements
    ]


@app.post("/parse-bank-statement")
async def parse_bank_statement(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """
    Parse a bank statement (CSV or PDF) and extract transactions.

    Returns list of transactions with date, description, and amount.
    If user is authenticated, saves to database for reconciliation history.
    """
    # Validate file upload
    await validate_file_upload(file)

    try:
        file_content = await file.read()
        file_type = file.content_type
        parsing_method = "basic"  # Track which parser was used

        # Initialize parser
        parser = BankStatementParser()

        # Parse the statement with basic parser first
        transactions = parser.parse(file_content, file_type)

        # If basic parser returns no transactions for PDF, try Gemini AI
        if len(transactions) == 0 and (file_type == 'pdf' or file_type == 'application/pdf'):
            print(f"Basic parser found 0 transactions in PDF, trying Gemini AI...")
            try:
                transactions = await parse_bank_statement_with_gemini(file_content)
                if len(transactions) > 0:
                    parsing_method = "gemini_ai"
                    print(f"Gemini AI successfully extracted {len(transactions)} transactions")
            except Exception as gemini_error:
                print(f"Gemini AI parsing failed: {str(gemini_error)}")
                # Keep transactions as empty list from basic parser

        # Save to database if user is authenticated
        db_statement = None
        if current_user:
            try:
                # Create bank statement record
                statement_data = {
                    "file_name": file.filename,
                    "file_type": file_type,
                    "transactions_data": transactions,
                    "transaction_count": len(transactions)
                }
                db_statement = crud.create_bank_statement(
                    db=db,
                    user_id=current_user.id,
                    statement_data=statement_data
                )

                # Save individual transactions
                for transaction in transactions:
                    transaction_data = {
                        "transaction_date": transaction.get("date"),
                        "description": transaction.get("description"),
                        "amount": transaction.get("amount"),
                        "transaction_type": transaction.get("type")
                    }
                    crud.create_bank_transaction(
                        db=db,
                        user_id=current_user.id,
                        bank_statement_id=db_statement.id,
                        transaction_data=transaction_data
                    )

                # Log activity
                crud.log_activity(
                    db=db,
                    user_id=current_user.id,
                    action="bank_statement_uploaded",
                    entity_type="bank_statement",
                    entity_id=db_statement.id,
                    details={"file_name": file.filename, "transaction_count": len(transactions)}
                )
            except Exception as e:
                print(f"Warning: Failed to save bank statement to database: {e}")

        return {
            "success": True,
            "transactions": transactions,
            "count": len(transactions),
            "file_name": file.filename,
            "file_type": file_type,
            "statement_id": db_statement.id if db_statement else None,
            "parsing_method": parsing_method  # Track which parser was used
        }

    except Exception as e:
        print(f"Error parsing bank statement: {str(e)}")
        return {
            "success": False,
            "error": f"Error parsing bank statement: {str(e)}"
        }

# Define request model for reconciliation
class ReconciliationRequest(BaseModel):
    documents: list  # List of processed documents
    bank_transactions: list  # List of bank transactions
    auto_match_threshold: int = 90  # Threshold for automatic matching

@app.post("/reconcile")
async def reconcile_documents(
    request: ReconciliationRequest,
    current_user: models.User = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """
    Reconcile documents against bank statement transactions.

    Uses intelligent matching based on:
    - Vendor name (fuzzy matching)
    - Amount (exact or within tolerance)
    - Date (exact or within range)

    Returns matched, unmatched, and suggested matches.
    If user is authenticated, saves matches to database for history.
    """
    try:
        # Initialize reconciliation engine
        engine = ReconciliationEngine(
            name_threshold=80,  # Minimum 80% name match
            amount_tolerance=0.01,  # Within $0.01
            date_range_days=3  # Within 3 days
        )

        # Perform reconciliation
        results = engine.reconcile(
            documents=request.documents,
            bank_transactions=request.bank_transactions,
            auto_match_threshold=request.auto_match_threshold
        )

        # Save matches to database if user is authenticated
        if current_user:
            try:
                for match in results.get("matched", []):
                    # Find transaction and bank transaction in database
                    doc_id = safe_get(match, "document", "id", default="")
                    bank_tx_id = safe_get(match, "bank_transaction", "id", default=None)
                    db_transaction = crud.get_transaction_by_id(
                        db, doc_id, current_user.id
                    )
                    db_bank_transaction = db.query(models.BankTransaction).filter(
                        models.BankTransaction.user_id == current_user.id,
                        models.BankTransaction.id == bank_tx_id
                    ).first() if bank_tx_id else None

                    if db_transaction and db_bank_transaction:
                        # Create reconciliation match
                        crud.create_reconciliation_match(
                            db=db,
                            user_id=current_user.id,
                            transaction_id=db_transaction.id,
                            bank_transaction_id=db_bank_transaction.id,
                            match_type="auto" if match.get("confidence", 0) >= request.auto_match_threshold else "suggested",
                            match_confidence=match.get("confidence"),
                            name_match_score=match.get("name_score"),
                            amount_match_score=match.get("amount_score"),
                            date_match_score=match.get("date_score"),
                            match_reason=match.get("reason"),
                            match_data=match
                        )

                # Log activity
                crud.log_activity(
                    db=db,
                    user_id=current_user.id,
                    action="reconciliation_performed",
                    entity_type="reconciliation",
                    details={
                        "matched_count": len(results.get("matched", [])),
                        "unmatched_count": len(results.get("unmatched", []))
                    }
                )
            except Exception as e:
                print(f"Warning: Failed to save reconciliation results to database: {e}")

        return {
            "success": True,
            "results": results
        }

    except Exception as e:
        print(f"Error during reconciliation: {str(e)}")
        return {
            "success": False,
            "error": f"Error during reconciliation: {str(e)}"
        }

# Define request model for manual matching
class ManualMatchRequest(BaseModel):
    document: dict
    transaction: dict

@app.post("/manual-match")
async def manual_match(request: ManualMatchRequest):
    """
    Manually match a document with a transaction.

    User confirms a suggested match or creates a new match.
    Returns detailed match information.
    """
    try:
        # Initialize reconciliation engine
        engine = ReconciliationEngine()

        # Create manual match
        match_result = engine.manual_match(
            document=request.document,
            transaction=request.transaction
        )

        return {
            "success": True,
            "match": match_result
        }

    except Exception as e:
        print(f"Error creating manual match: {str(e)}")
        return {
            "success": False,
            "error": f"Error creating manual match: {str(e)}"
        }

# ============================================================================
# AUTHENTICATION ENDPOINTS
# ============================================================================

class UserCreate(BaseModel):
    username: str
    email: EmailStr
    password: str


class UserResponse(BaseModel):
    id: int
    username: str
    email: str
    role: str
    is_active: bool
    created_at: datetime

    class Config:
        from_attributes = True


class Token(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse


@app.post("/register", response_model=UserResponse, tags=["Authentication"])
async def register(user_data: UserCreate, db: Session = Depends(get_db)):
    """
    Register a new user

    - **username**: Unique username
    - **email**: Unique email address
    - **password**: Password (will be hashed)
    """
    # Check if username exists
    existing_user = crud.get_user_by_username(db, user_data.username)
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Username already registered"
        )

    # Check if email exists
    existing_email = crud.get_user_by_email(db, user_data.email)
    if existing_email:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email already registered"
        )

    # Create user
    user = crud.create_user(
        db=db,
        username=user_data.username,
        email=user_data.email,
        password=user_data.password
    )

    # Log activity
    crud.log_activity(
        db=db,
        user_id=user.id,
        action="user_registered",
        entity_type="user",
        entity_id=user.id,
        details={"username": user.username, "email": user.email}
    )

    return user


@app.post("/login", response_model=Token, tags=["Authentication"])
async def login(
    form_data: OAuth2PasswordRequestForm = Depends(),
    db: Session = Depends(get_db)
):
    """
    Login with username and password

    Returns JWT access token for authentication
    """
    user = authenticate_user(form_data.username, form_data.password, db)

    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )

    # Update last login
    crud.update_user_login(db, user.id)

    # Create access token
    access_token = create_access_token(data={"sub": user.username})

    # Log activity
    crud.log_activity(
        db=db,
        user_id=user.id,
        action="user_login",
        entity_type="user",
        entity_id=user.id
    )

    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": user
    }


@app.get("/me", response_model=UserResponse, tags=["Authentication"])
async def get_current_user_info(current_user: models.User = Depends(get_current_user)):
    """Get current authenticated user information"""
    return current_user


# ============================================================================
# USER SETTINGS ENDPOINTS
# ============================================================================

class UserSettings(BaseModel):
    """User preferences and settings"""
    confidence_threshold: float = Field(
        default=70.0,
        ge=0,
        le=100,
        description="Minimum confidence score (0-100) to auto-approve categorizations"
    )
    auto_approve_vendor_mapping: bool = Field(
        default=True,
        description="Automatically approve categorizations from known vendor mapping"
    )
    default_export_format: str = Field(
        default="csv",
        description="Default export format (csv or excel)"
    )

    class Config:
        from_attributes = True


class UserSettingsResponse(BaseModel):
    """Response model for user settings"""
    settings: UserSettings
    message: str


DEFAULT_USER_SETTINGS = {
    "confidence_threshold": 70.0,
    "auto_approve_vendor_mapping": True,
    "default_export_format": "csv"
}


@app.get("/settings", response_model=UserSettingsResponse, tags=["User Settings"])
async def get_user_settings(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Get current user's settings/preferences.

    Returns the user's configured settings including:
    - **confidence_threshold**: Minimum confidence for auto-approval (default: 70%)
    - **auto_approve_vendor_mapping**: Auto-approve known vendors (default: true)
    - **default_export_format**: Preferred export format (default: csv)
    """
    # Merge user settings with defaults
    user_settings = current_user.settings or {}
    merged_settings = {**DEFAULT_USER_SETTINGS, **user_settings}

    return {
        "settings": merged_settings,
        "message": "Settings retrieved successfully"
    }


@app.patch("/settings", response_model=UserSettingsResponse, tags=["User Settings"])
async def update_user_settings(
    settings: UserSettings,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Update current user's settings/preferences.

    Configurable settings:
    - **confidence_threshold**: Set between 0-100 (higher = stricter auto-approval)
    - **auto_approve_vendor_mapping**: Enable/disable auto-approval for known vendors
    - **default_export_format**: Set to 'csv' or 'excel'
    """
    # Get existing settings or empty dict
    existing_settings = current_user.settings or {}

    # Update with new values
    updated_settings = {
        **existing_settings,
        "confidence_threshold": settings.confidence_threshold,
        "auto_approve_vendor_mapping": settings.auto_approve_vendor_mapping,
        "default_export_format": settings.default_export_format
    }

    # Save to database
    current_user.settings = updated_settings
    db.commit()
    db.refresh(current_user)

    return {
        "settings": updated_settings,
        "message": "Settings updated successfully"
    }


@app.patch("/settings/confidence-threshold", response_model=UserSettingsResponse, tags=["User Settings"])
async def update_confidence_threshold(
    threshold: float = Query(..., ge=0, le=100, description="New confidence threshold (0-100)"),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Quick endpoint to update just the confidence threshold.

    - **threshold**: Value between 0 and 100
      - 0-50: Very lenient (most categorizations auto-approved)
      - 50-70: Moderate (good balance)
      - 70-85: Strict (only high-confidence auto-approved)
      - 85-100: Very strict (manual review for most)
    """
    existing_settings = current_user.settings or {}
    existing_settings["confidence_threshold"] = threshold

    current_user.settings = existing_settings
    db.commit()
    db.refresh(current_user)

    merged_settings = {**DEFAULT_USER_SETTINGS, **existing_settings}

    return {
        "settings": merged_settings,
        "message": f"Confidence threshold updated to {threshold}%"
    }


# ============================================================================
# DOCUMENT & TRANSACTION HISTORY ENDPOINTS
# ============================================================================

class DocumentResponse(BaseModel):
    id: int
    document_id: str
    file_name: str
    status: str
    progress: int
    uploaded_at: datetime
    processed_at: Optional[datetime]

    class Config:
        from_attributes = True


class TransactionResponse(BaseModel):
    id: int
    transaction_id: str
    vendor_name: Optional[str]
    amount: float
    transaction_date: Optional[date]
    description: Optional[str]
    created_at: datetime

    class Config:
        from_attributes = True


@app.get("/documents", response_model=List[DocumentResponse], tags=["Documents"])
async def get_documents(
    skip: int = 0,
    limit: int = 100,
    status: Optional[str] = None,
    current_user: models.User = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """
    Get all documents for current user (or all if not authenticated)

    - **skip**: Number of records to skip (pagination)
    - **limit**: Maximum number of records to return
    - **status**: Filter by status (pending, processing, completed, error)
    """
    if current_user:
        documents = crud.get_user_documents(
            db=db,
            user_id=current_user.id,
            skip=skip,
            limit=limit,
            status=status
        )
    else:
        # For non-authenticated users, return empty list or demo data
        documents = []

    return documents


@app.get("/documents/{document_id}", response_model=DocumentResponse, tags=["Documents"])
async def get_document(
    document_id: str,
    current_user: models.User = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """Get a specific document by ID"""
    if not current_user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required"
        )

    document = crud.get_document_by_id(db, document_id, current_user.id)

    if not document:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Document not found"
        )

    return document


class TransactionSearchRequest(BaseModel):
    search_query: Optional[str] = None
    vendor_name: Optional[str] = None
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    min_amount: Optional[float] = None
    max_amount: Optional[float] = None
    category: Optional[str] = None
    skip: int = 0
    limit: int = 100


@app.post("/transactions/search", response_model=List[TransactionResponse], tags=["Transactions"])
async def search_transactions(
    search_request: TransactionSearchRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Search and filter transactions

    Supports:
    - Full-text search
    - Vendor name filtering
    - Date range filtering
    - Amount range filtering
    - Category filtering
    """
    if search_request.search_query:
        # Full-text search
        transactions = crud.search_transactions(
            db=db,
            user_id=current_user.id,
            search_query=search_request.search_query,
            skip=search_request.skip,
            limit=search_request.limit
        )
    else:
        # Filtered search
        transactions = crud.get_user_transactions(
            db=db,
            user_id=current_user.id,
            vendor_name=search_request.vendor_name,
            start_date=search_request.start_date,
            end_date=search_request.end_date,
            min_amount=search_request.min_amount,
            max_amount=search_request.max_amount,
            category=search_request.category,
            skip=search_request.skip,
            limit=search_request.limit
        )

    return transactions


@app.get("/transactions", response_model=List[TransactionResponse], tags=["Transactions"])
async def get_transactions(
    skip: int = 0,
    limit: int = 100,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all transactions for current user"""
    transactions = crud.get_user_transactions(
        db=db,
        user_id=current_user.id,
        skip=skip,
        limit=limit
    )
    return transactions


@app.delete("/documents/{document_id}", tags=["Documents"])
async def delete_document(
    document_id: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a document (and all associated transactions)"""
    crud.delete_document(db, document_id, current_user.id)

    # Log activity
    crud.log_activity(
        db=db,
        user_id=current_user.id,
        action="document_deleted",
        entity_type="document",
        details={"document_id": document_id}
    )

    return {"success": True, "message": "Document deleted"}


# ============================================================================
# MANUAL REVIEW QUEUE ENDPOINTS
# ============================================================================

class LowConfidenceTransaction(BaseModel):
    """Extended transaction response with confidence and review info"""
    id: int
    transaction_id: str
    vendor_name: Optional[str]
    amount: float
    transaction_date: Optional[date]
    description: Optional[str]
    created_at: datetime

    # Categorization info
    category: Optional[str]
    subcategory: Optional[str]
    confidence_score: Optional[float]
    needs_review_reason: Optional[str]

    # Research info
    has_research: bool
    research_confidence: Optional[float]

    class Config:
        from_attributes = True


@app.get("/review-queue", response_model=List[dict], tags=["Review"])
async def get_review_queue(
    skip: int = 0,
    limit: int = 50,
    min_confidence: Optional[float] = None,
    max_confidence: Optional[float] = 70.0,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Get transactions flagged for manual review.

    Returns transactions where:
    - Confidence score is below threshold (default < 70%)
    - Notes contain "NEEDS REVIEW"
    - Categorization is not user-approved

    Sorted by confidence score (lowest first).
    """
    # Query transactions with low confidence categorizations
    query = db.query(
        models.Transaction,
        models.Categorization
    ).join(
        models.Categorization,
        models.Transaction.id == models.Categorization.transaction_id,
        isouter=True
    ).filter(
        models.Transaction.user_id == current_user.id
    ).filter(
        or_(
            # Low confidence score
            models.Categorization.confidence_score < max_confidence if max_confidence else True,
            # Not user approved
            models.Categorization.user_approved == False,
            # Flagged in notes
            models.Transaction.notes.like("%NEEDS REVIEW%")
        )
    )

    if min_confidence is not None:
        query = query.filter(models.Categorization.confidence_score >= min_confidence)

    # Order by confidence (lowest first)
    query = query.order_by(
        models.Categorization.confidence_score.asc().nullsfirst()
    )

    results = query.offset(skip).limit(limit).all()

    # Format results
    review_queue = []
    for transaction, categorization in results:
        item = {
            "id": transaction.id,
            "transaction_id": transaction.transaction_id,
            "vendor_name": transaction.vendor_name,
            "amount": float(transaction.amount) if transaction.amount else 0,
            "transaction_date": transaction.transaction_date,
            "description": transaction.description,
            "created_at": transaction.created_at,
            "category": categorization.category if categorization else None,
            "subcategory": categorization.subcategory if categorization else None,
            "confidence_score": float(categorization.confidence_score) if categorization and categorization.confidence_score else 0,
            "needs_review_reason": transaction.notes if "NEEDS REVIEW" in (transaction.notes or "") else "Low confidence score",
            "categorization_method": categorization.method if categorization else None,
            "user_approved": categorization.user_approved if categorization else False
        }
        review_queue.append(item)

    return review_queue


@app.get("/review-queue/stats", tags=["Review"])
async def get_review_queue_stats(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Get statistics about the review queue.

    Returns counts by confidence range and urgency level.
    """
    # Total transactions needing review
    total_needs_review = db.query(models.Transaction).join(
        models.Categorization
    ).filter(
        models.Transaction.user_id == current_user.id
    ).filter(
        or_(
            models.Categorization.confidence_score < 70,
            models.Categorization.user_approved == False,
            models.Transaction.notes.like("%NEEDS REVIEW%")
        )
    ).count()

    # Count by confidence ranges
    critical = db.query(models.Transaction).join(
        models.Categorization
    ).filter(
        models.Transaction.user_id == current_user.id,
        models.Categorization.confidence_score < 50
    ).count()

    low = db.query(models.Transaction).join(
        models.Categorization
    ).filter(
        models.Transaction.user_id == current_user.id,
        models.Categorization.confidence_score >= 50,
        models.Categorization.confidence_score < 70
    ).count()

    # Recently added (last 7 days)
    from datetime import timedelta
    recent_date = datetime.utcnow() - timedelta(days=7)
    recent = db.query(models.Transaction).join(
        models.Categorization
    ).filter(
        models.Transaction.user_id == current_user.id,
        models.Transaction.created_at >= recent_date,
        models.Categorization.confidence_score < 70
    ).count()

    return {
        "total_needs_review": total_needs_review,
        "by_urgency": {
            "critical": critical,  # < 50%
            "low": low,  # 50-70%
        },
        "recent_additions": recent,
        "average_queue_time_days": 0,  # Could calculate based on created_at
        "oldest_unreviewed": None  # Could query for oldest transaction
    }


class ApproveCategorizationRequest(BaseModel):
    transaction_id: str
    approved: bool
    corrected_category: Optional[str] = None
    corrected_subcategory: Optional[str] = None
    corrected_ledger_type: Optional[str] = None
    review_notes: Optional[str] = None


class ApproveBankTransactionRequest(BaseModel):
    """Request model for approving/correcting a bank transaction categorization"""
    bank_transaction_id: int
    approved: bool
    corrected_category: Optional[str] = None
    corrected_subcategory: Optional[str] = None
    corrected_ledger_type: Optional[str] = None
    review_notes: Optional[str] = None


class BulkApproveBankTransactionsRequest(BaseModel):
    """Request model for bulk approving bank transaction categorizations"""
    bank_transaction_ids: Optional[List[int]] = Field(
        default=None,
        description="List of specific bank transaction IDs to approve"
    )
    bank_statement_id: Optional[int] = Field(
        default=None,
        description="Approve all categorized transactions for this bank statement"
    )
    min_confidence: Optional[float] = Field(
        default=None,
        ge=0,
        le=100,
        description="Only approve transactions with confidence >= this threshold"
    )
    approve_all_high_confidence: bool = Field(
        default=False,
        description="Approve all high-confidence transactions (uses user's confidence threshold)"
    )


@app.post("/bank-transaction/bulk-approve", tags=["Bank Statements"])
async def bulk_approve_bank_transaction_categorizations(
    request: BulkApproveBankTransactionsRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Bulk approve multiple bank transaction categorizations at once.

    **Options:**
    1. **By transaction IDs**: Provide a list of specific bank_transaction_ids to approve
    2. **By bank statement**: Provide bank_statement_id to approve all categorized transactions
    3. **By confidence threshold**: Add min_confidence to only approve above that confidence
    4. **High confidence auto-approve**: Set approve_all_high_confidence=true
    """
    try:
        # Validate request
        if not request.bank_transaction_ids and not request.bank_statement_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Must provide either bank_transaction_ids or bank_statement_id"
            )

        # Get user's confidence threshold
        user_threshold = 70.0
        if request.approve_all_high_confidence:
            user_settings = current_user.settings or {}
            user_threshold = user_settings.get("confidence_threshold", 70.0)

        # Determine effective threshold
        effective_threshold = None
        if request.approve_all_high_confidence:
            effective_threshold = user_threshold
        elif request.min_confidence is not None:
            effective_threshold = request.min_confidence

        approved_count = 0
        skipped_count = 0
        approved_ids = []
        skipped_info = []

        # Get bank transactions to process
        if request.bank_transaction_ids:
            bank_txs = db.query(models.BankTransaction).filter(
                models.BankTransaction.id.in_(request.bank_transaction_ids),
                models.BankTransaction.user_id == current_user.id
            ).all()
        else:
            # Verify statement exists
            statement = crud.get_bank_statement_by_id(db, current_user.id, request.bank_statement_id)
            if not statement:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Bank statement not found"
                )
            bank_txs = crud.get_bank_transactions_by_statement(db, current_user.id, request.bank_statement_id)

        for bank_tx in bank_txs:
            # Get categorization
            categorization = crud.get_categorization_for_bank_transaction(
                db, current_user.id, bank_tx.id
            )

            if not categorization:
                skipped_info.append({"id": bank_tx.id, "reason": "No categorization"})
                skipped_count += 1
                continue

            if categorization.user_approved:
                skipped_info.append({"id": bank_tx.id, "reason": "Already approved"})
                skipped_count += 1
                continue

            # Check confidence threshold
            if effective_threshold is not None:
                confidence = float(categorization.confidence_score) if categorization.confidence_score else 0
                if confidence < effective_threshold:
                    skipped_info.append({
                        "id": bank_tx.id,
                        "reason": f"Below threshold ({confidence:.1f}% < {effective_threshold}%)"
                    })
                    skipped_count += 1
                    continue

            # Approve
            categorization.user_approved = True
            categorization.user_modified = False
            approved_ids.append(bank_tx.id)
            approved_count += 1

        db.commit()

        # Log activity
        crud.log_activity(
            db=db,
            user_id=current_user.id,
            action="bulk_bank_categorization_approved",
            entity_type="bank_statement",
            entity_id=request.bank_statement_id,
            details={
                "approved_count": approved_count,
                "skipped_count": skipped_count,
                "threshold": effective_threshold
            }
        )

        return {
            "success": True,
            "approved_count": approved_count,
            "skipped_count": skipped_count,
            "total_processed": approved_count + skipped_count,
            "message": f"Approved {approved_count} transactions, skipped {skipped_count}",
            "approved_transactions": approved_ids,
            "skipped_transactions": skipped_info
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error in bulk approve: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error during bulk approve: {str(e)}"
        )


@app.post("/bank-transaction/approve", tags=["Bank Statements"])
async def approve_bank_transaction_categorization(
    request: ApproveBankTransactionRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Approve or correct a bank transaction categorization.

    If approved: Mark categorization as user_approved
    If corrected: Update categorization with new values
    """
    try:
        # Find the bank transaction
        bank_tx = db.query(models.BankTransaction).filter(
            models.BankTransaction.id == request.bank_transaction_id,
            models.BankTransaction.user_id == current_user.id
        ).first()

        if not bank_tx:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Bank transaction not found"
            )

        # Get categorization for this bank transaction
        categorization = crud.get_categorization_for_bank_transaction(
            db, current_user.id, request.bank_transaction_id
        )

        if not categorization:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Categorization not found for this transaction"
            )

        if request.approved:
            # Simple approval
            categorization.user_approved = True
            categorization.user_modified = False
            db.commit()

            # Log activity
            crud.log_activity(
                db=db,
                user_id=current_user.id,
                action="bank_categorization_approved",
                entity_type="categorization",
                entity_id=categorization.id,
                details={"bank_transaction_id": request.bank_transaction_id}
            )

            return {
                "success": True,
                "message": "Categorization approved",
                "action": "approved"
            }

        else:
            # User is correcting the categorization
            if not request.corrected_category:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Corrected category is required"
                )

            # Update the categorization
            categorization.category = request.corrected_category
            categorization.subcategory = request.corrected_subcategory or ""
            categorization.ledger_type = request.corrected_ledger_type or "Expense"
            categorization.method = "manual"
            categorization.confidence_score = 100.0  # User corrections are 100% confident
            categorization.user_approved = True
            categorization.user_modified = True

            # Also update the bank transaction category field
            bank_tx.category = request.corrected_category

            db.commit()

            # Log activity
            crud.log_activity(
                db=db,
                user_id=current_user.id,
                action="bank_categorization_corrected",
                entity_type="categorization",
                entity_id=categorization.id,
                details={
                    "bank_transaction_id": request.bank_transaction_id,
                    "corrected_category": request.corrected_category,
                    "corrected_subcategory": request.corrected_subcategory
                }
            )

            return {
                "success": True,
                "message": "Categorization corrected",
                "action": "corrected"
            }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error approving bank transaction categorization: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing approval: {str(e)}"
        )


@app.post("/review-queue/approve", tags=["Review"])
async def approve_categorization(
    request: ApproveCategorizationRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Approve or correct a categorization from the review queue.

    If approved: Mark categorization as user_approved
    If corrected: Create user correction and update categorization
    """
    try:
        # Find transaction
        transaction = crud.get_transaction_by_id(db, request.transaction_id, current_user.id)

        if not transaction:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Transaction not found"
            )

        # Get current categorization
        categorization = db.query(models.Categorization).filter(
            models.Categorization.transaction_id == transaction.id,
            models.Categorization.user_id == current_user.id
        ).order_by(models.Categorization.created_at.desc()).first()

        if not categorization:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Categorization not found"
            )

        if request.approved:
            # Simple approval
            categorization.user_approved = True
            categorization.user_modified = False

            # Clear review flag from notes
            if transaction.notes and "NEEDS REVIEW" in transaction.notes:
                transaction.notes = transaction.notes.replace("NEEDS REVIEW - ", "").strip()

            db.commit()

            # Log activity
            crud.log_activity(
                db=db,
                user_id=current_user.id,
                action="categorization_approved",
                entity_type="categorization",
                entity_id=categorization.id,
                details={"transaction_id": request.transaction_id}
            )

            return {
                "success": True,
                "message": "Categorization approved",
                "action": "approved"
            }

        else:
            # User is correcting the categorization
            if not (request.corrected_category and request.corrected_subcategory and request.corrected_ledger_type):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Corrected category, subcategory, and ledger type required"
                )

            # Create user correction record
            crud.create_user_correction(
                db=db,
                user_id=current_user.id,
                transaction_id=transaction.id,
                original_category=categorization.category,
                original_subcategory=categorization.subcategory,
                original_ledger_type=categorization.ledger_type,
                original_method=categorization.method,
                corrected_category=request.corrected_category,
                corrected_subcategory=request.corrected_subcategory,
                corrected_ledger_type=request.corrected_ledger_type,
                correction_reason=request.review_notes or "Manual review correction",
                correction_data={
                    "original_confidence": float(categorization.confidence_score) if categorization.confidence_score else 0
                }
            )

            # Update categorization
            categorization.category = request.corrected_category
            categorization.subcategory = request.corrected_subcategory
            categorization.ledger_type = request.corrected_ledger_type
            categorization.user_approved = True
            categorization.user_modified = True
            categorization.confidence_score = 100  # User correction is 100% confident

            # Clear review flag
            if transaction.notes and "NEEDS REVIEW" in transaction.notes:
                transaction.notes = f"Manually corrected: {request.review_notes or 'User correction'}"

            db.commit()

            # Log activity
            crud.log_activity(
                db=db,
                user_id=current_user.id,
                action="categorization_corrected",
                entity_type="categorization",
                entity_id=categorization.id,
                details={
                    "transaction_id": request.transaction_id,
                    "corrected_to": request.corrected_category
                }
            )

            return {
                "success": True,
                "message": "Categorization corrected",
                "action": "corrected"
            }

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error approving categorization: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing approval: {str(e)}"
        )


# ============================================================================
# BULK APPROVE ENDPOINT
# ============================================================================

class BulkApproveRequest(BaseModel):
    """Request model for bulk approving categorizations"""
    transaction_ids: Optional[List[str]] = Field(
        default=None,
        description="List of specific transaction IDs to approve"
    )
    bank_statement_id: Optional[int] = Field(
        default=None,
        description="Approve all transactions for this bank statement"
    )
    min_confidence: Optional[float] = Field(
        default=None,
        ge=0,
        le=100,
        description="Only approve transactions with confidence >= this threshold"
    )
    approve_all_high_confidence: bool = Field(
        default=False,
        description="Approve all high-confidence transactions (uses user's confidence threshold)"
    )


class BulkApproveResponse(BaseModel):
    """Response model for bulk approve operation"""
    success: bool
    approved_count: int
    skipped_count: int
    total_processed: int
    message: str
    approved_transactions: List[str] = []
    skipped_transactions: List[dict] = []


@app.post("/review-queue/bulk-approve", tags=["Review"], response_model=BulkApproveResponse)
async def bulk_approve_categorizations(
    request: BulkApproveRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Bulk approve multiple categorizations at once.

    **Options:**
    1. **By transaction IDs**: Provide a list of specific transaction_ids to approve
    2. **By bank statement**: Provide bank_statement_id to approve all categorized transactions for that statement
    3. **By confidence threshold**: Add min_confidence to only approve transactions above that confidence
    4. **High confidence auto-approve**: Set approve_all_high_confidence=true to approve all transactions
       meeting the user's confidence threshold setting

    **Examples:**
    - Approve specific transactions: `{"transaction_ids": ["tx1", "tx2", "tx3"]}`
    - Approve all for a statement: `{"bank_statement_id": 1}`
    - Approve high-confidence only: `{"bank_statement_id": 1, "min_confidence": 85}`
    - Auto-approve by user threshold: `{"bank_statement_id": 1, "approve_all_high_confidence": true}`

    **Note:** Already approved transactions will be skipped.
    """
    try:
        # Validate request - need at least one selection criteria
        if not request.transaction_ids and not request.bank_statement_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Must provide either transaction_ids or bank_statement_id"
            )

        # Get user's confidence threshold if using auto-approve
        user_threshold = 70.0  # default
        if request.approve_all_high_confidence:
            user_settings = crud.get_user_settings(db, current_user.id)
            if user_settings and user_settings.settings:
                user_threshold = user_settings.settings.get("confidence_threshold", 70.0)

        # Determine effective confidence threshold
        effective_threshold = None
        if request.approve_all_high_confidence:
            effective_threshold = user_threshold
        elif request.min_confidence is not None:
            effective_threshold = request.min_confidence

        approved_count = 0
        skipped_count = 0
        approved_transactions = []
        skipped_transactions = []

        # Build query for categorizations to approve
        if request.transaction_ids:
            # Get transactions by IDs
            transactions = db.query(models.Transaction).filter(
                models.Transaction.transaction_id.in_(request.transaction_ids),
                models.Transaction.user_id == current_user.id
            ).all()

            transaction_id_map = {t.transaction_id: t.id for t in transactions}

            for tx_id in request.transaction_ids:
                if tx_id not in transaction_id_map:
                    skipped_transactions.append({
                        "transaction_id": tx_id,
                        "reason": "Transaction not found"
                    })
                    skipped_count += 1
                    continue

                # Get categorization
                categorization = db.query(models.Categorization).filter(
                    models.Categorization.transaction_id == transaction_id_map[tx_id],
                    models.Categorization.user_id == current_user.id
                ).order_by(models.Categorization.created_at.desc()).first()

                if not categorization:
                    skipped_transactions.append({
                        "transaction_id": tx_id,
                        "reason": "No categorization found"
                    })
                    skipped_count += 1
                    continue

                if categorization.user_approved:
                    skipped_transactions.append({
                        "transaction_id": tx_id,
                        "reason": "Already approved"
                    })
                    skipped_count += 1
                    continue

                # Check confidence threshold
                if effective_threshold is not None:
                    confidence = float(categorization.confidence_score) if categorization.confidence_score else 0
                    if confidence < effective_threshold:
                        skipped_transactions.append({
                            "transaction_id": tx_id,
                            "reason": f"Below confidence threshold ({confidence:.1f}% < {effective_threshold}%)"
                        })
                        skipped_count += 1
                        continue

                # Approve the categorization
                categorization.user_approved = True
                categorization.user_modified = False
                approved_transactions.append(tx_id)
                approved_count += 1

        elif request.bank_statement_id:
            # Verify bank statement exists and belongs to user
            statement = crud.get_bank_statement_by_id(db, current_user.id, request.bank_statement_id)
            if not statement:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Bank statement not found"
                )

            # Get all bank transactions for this statement
            bank_txs = crud.get_bank_transactions_by_statement(db, current_user.id, request.bank_statement_id)

            for bank_tx in bank_txs:
                tx_id = bank_tx.transaction_id if hasattr(bank_tx, 'transaction_id') else None

                if not tx_id:
                    skipped_count += 1
                    continue

                # Get the transaction record
                transaction = db.query(models.Transaction).filter(
                    models.Transaction.id == tx_id,
                    models.Transaction.user_id == current_user.id
                ).first()

                if not transaction:
                    skipped_count += 1
                    continue

                # Get categorization
                categorization = db.query(models.Categorization).filter(
                    models.Categorization.transaction_id == transaction.id,
                    models.Categorization.user_id == current_user.id
                ).order_by(models.Categorization.created_at.desc()).first()

                if not categorization:
                    skipped_transactions.append({
                        "transaction_id": transaction.transaction_id,
                        "reason": "No categorization found"
                    })
                    skipped_count += 1
                    continue

                if categorization.user_approved:
                    skipped_transactions.append({
                        "transaction_id": transaction.transaction_id,
                        "reason": "Already approved"
                    })
                    skipped_count += 1
                    continue

                # Check confidence threshold
                if effective_threshold is not None:
                    confidence = float(categorization.confidence_score) if categorization.confidence_score else 0
                    if confidence < effective_threshold:
                        skipped_transactions.append({
                            "transaction_id": transaction.transaction_id,
                            "reason": f"Below confidence threshold ({confidence:.1f}% < {effective_threshold}%)"
                        })
                        skipped_count += 1
                        continue

                # Approve the categorization
                categorization.user_approved = True
                categorization.user_modified = False
                approved_transactions.append(transaction.transaction_id)
                approved_count += 1

        # Commit all approvals
        db.commit()

        # Log activity
        crud.log_activity(
            db=db,
            user_id=current_user.id,
            action="bulk_approve",
            entity_type="categorizations",
            entity_id=None,
            details={
                "approved_count": approved_count,
                "skipped_count": skipped_count,
                "bank_statement_id": request.bank_statement_id,
                "min_confidence": effective_threshold
            }
        )

        total_processed = approved_count + skipped_count

        return BulkApproveResponse(
            success=True,
            approved_count=approved_count,
            skipped_count=skipped_count,
            total_processed=total_processed,
            message=f"Approved {approved_count} of {total_processed} transactions",
            approved_transactions=approved_transactions,
            skipped_transactions=skipped_transactions[:50]  # Limit to first 50 for response size
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in bulk approve: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing bulk approval: {str(e)}"
        )


# ============================================================================
# BATCH CATEGORIZATION ENDPOINTS
# ============================================================================

class BatchCategorizationRequest(BaseModel):
    """Request model for batch categorization of bank statement transactions"""
    bank_statement_id: int
    confidence_threshold: Optional[float] = Field(
        default=None,
        ge=0,
        le=100,
        description="Minimum confidence to auto-approve. If not provided, uses user's saved setting (default: 70%)"
    )
    use_vendor_mapping: Optional[bool] = Field(
        default=None,
        description="Use vendor mapping for known merchants. If not provided, uses user's saved setting (default: true)"
    )


class BatchCategorizationResponse(BaseModel):
    """Response model for batch categorization results"""
    success: bool
    statement_id: int
    total_transactions: int
    processed: int
    failed: int
    high_confidence: int
    low_confidence: int
    results: List[dict]
    summary: dict


@app.post("/categorize-bank-statement", response_model=BatchCategorizationResponse)
async def categorize_bank_statement(
    request: BatchCategorizationRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Batch categorize ALL transactions in a bank statement.

    This endpoint:
    1. Retrieves all transactions from the specified bank statement
    2. Categorizes each transaction using the hybrid ML + Gemini approach
    3. Stores categorizations in the database
    4. Returns comprehensive results with confidence scores

    Transactions with confidence >= threshold are marked as auto-approved.
    Low-confidence transactions are flagged for manual review.
    """
    # Verify the bank statement exists and belongs to the user
    statement = crud.get_bank_statement_by_id(db, current_user.id, request.bank_statement_id)
    if not statement:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Bank statement with ID {request.bank_statement_id} not found"
        )

    # Get user settings and resolve threshold/options
    user_settings = current_user.settings or {}
    confidence_threshold = request.confidence_threshold
    if confidence_threshold is None:
        confidence_threshold = user_settings.get("confidence_threshold", DEFAULT_USER_SETTINGS["confidence_threshold"])

    use_vendor_mapping = request.use_vendor_mapping
    if use_vendor_mapping is None:
        use_vendor_mapping = user_settings.get("auto_approve_vendor_mapping", DEFAULT_USER_SETTINGS["auto_approve_vendor_mapping"])

    # Get all transactions for this statement
    bank_transactions = crud.get_bank_transactions_by_statement(
        db, current_user.id, request.bank_statement_id
    )

    if not bank_transactions:
        return BatchCategorizationResponse(
            success=True,
            statement_id=request.bank_statement_id,
            total_transactions=0,
            processed=0,
            failed=0,
            high_confidence=0,
            low_confidence=0,
            results=[],
            summary={"message": "No transactions found in this statement"}
        )

    results = []
    processed = 0
    failed = 0
    high_confidence = 0
    low_confidence = 0
    category_counts = {}

    # Process each transaction
    for bank_tx in bank_transactions:
        try:
            # Check if already categorized
            existing_cat = crud.get_categorization_for_bank_transaction(
                db, current_user.id, bank_tx.id
            )
            if existing_cat:
                # Already categorized, include in results but skip processing
                results.append({
                    "bank_transaction_id": bank_tx.id,
                    "description": bank_tx.description,
                    "amount": float(bank_tx.amount),
                    "date": str(bank_tx.transaction_date) if bank_tx.transaction_date else None,
                    "category": existing_cat.category,
                    "subcategory": existing_cat.subcategory,
                    "ledger_type": existing_cat.ledger_type,
                    "confidence": float(existing_cat.confidence_score) if existing_cat.confidence_score else 0,
                    "method": existing_cat.method,
                    "status": "already_categorized",
                    "user_approved": existing_cat.user_approved
                })
                processed += 1
                if existing_cat.confidence_score and existing_cat.confidence_score >= confidence_threshold:
                    high_confidence += 1
                else:
                    low_confidence += 1
                category_counts[existing_cat.category] = category_counts.get(existing_cat.category, 0) + 1
                continue

            # Build document data structure for categorization
            document_data = {
                "documentMetadata": {
                    "source": {"name": bank_tx.description},
                    "documentType": "bank_transaction",
                    "documentDate": str(bank_tx.transaction_date) if bank_tx.transaction_date else None,
                    "documentNumber": f"bank_tx_{bank_tx.id}"
                },
                "financialData": {
                    "totalAmount": float(bank_tx.amount),
                    "currency": "USD",
                    "transactionType": bank_tx.transaction_type or ("debit" if bank_tx.amount < 0 else "credit")
                }
            }

            # STEP 1: Try vendor mapping first (fast, deterministic) if enabled
            vendor_result = None
            if use_vendor_mapping:
                vendor_result = categorize_by_vendor(bank_tx.description)

            if vendor_result:
                # Known vendor found - use deterministic categorization
                category = vendor_result["category"]
                subcategory = vendor_result["subcategory"]
                ledger_type = vendor_result["ledger_type"]
                confidence = vendor_result["confidence"]
                method = "vendor_mapping"
                explanation = vendor_result["explanation"]
                ml_conf = 0
                gemini_conf = 0
            else:
                # STEP 2: Fall back to ML + Gemini for unknown vendors
                try:
                    engine = get_ml_categorization_engine()

                    # Run ML prediction and Gemini categorization in parallel
                    ml_prediction_task = engine.predict_category(
                        document_data,
                        "Bank statement transaction"
                    )
                    gemini_task = _get_gemini_categorization(
                        bank_tx.description,
                        document_data,
                        "Bank statement transaction"
                    )

                    ml_prediction, gemini_result = await asyncio.gather(
                        ml_prediction_task,
                        gemini_task
                    )

                    # Determine best result (prefer higher confidence)
                    ml_conf = ml_prediction.get("confidence", 0) * 100 if ml_prediction.get("hasPrediction") else 0
                    gemini_conf = gemini_result.get("confidence", 0)

                    if ml_conf > gemini_conf and ml_prediction.get("hasPrediction"):
                        category = ml_prediction.get("category", "Other Expenses")
                        subcategory = ml_prediction.get("subcategory", "Miscellaneous")
                        ledger_type = ml_prediction.get("ledgerType", "Expense (Other)")
                        confidence = ml_conf
                        method = "ml"
                        explanation = f"ML prediction based on {ml_prediction.get('supportingTransactions', 0)} similar transactions"
                    else:
                        category = gemini_result.get("category", "Other Expenses")
                        subcategory = gemini_result.get("subcategory", "Miscellaneous")
                        ledger_type = gemini_result.get("ledgerType", "Expense (Other)")
                        confidence = gemini_conf
                        method = "gemini"
                        explanation = gemini_result.get("explanation", "AI categorization")

                    # Use hybrid if both have predictions
                    if ml_prediction.get("hasPrediction") and gemini_conf > 0:
                        method = "hybrid"
                        # Average confidence if both agree on category
                        if ml_prediction.get("category") == gemini_result.get("category"):
                            confidence = (ml_conf + gemini_conf) / 2

                except ValueError:
                    # ML engine not available, use Gemini only
                    gemini_result = await _get_gemini_categorization(
                        bank_tx.description,
                        document_data,
                        "Bank statement transaction"
                    )
                    category = gemini_result.get("category", "Other Expenses")
                    subcategory = gemini_result.get("subcategory", "Miscellaneous")
                    ledger_type = gemini_result.get("ledgerType", "Expense (Other)")
                    confidence = gemini_result.get("confidence", 0)
                    method = "gemini"
                    explanation = gemini_result.get("explanation", "AI categorization")
                    ml_conf = 0
                    gemini_conf = confidence

            # Determine if auto-approved based on confidence threshold
            auto_approved = confidence >= confidence_threshold

            # Create categorization record
            categorization_data = {
                "category": category,
                "subcategory": subcategory,
                "ledger_type": ledger_type,
                "method": method,
                "confidence_score": confidence,
                "ml_confidence": ml_conf,
                "gemini_confidence": gemini_conf,
                "explanation": explanation,
                "transaction_purpose": "Bank statement transaction"
            }

            db_categorization = crud.create_categorization(
                db=db,
                user_id=current_user.id,
                categorization_data=categorization_data,
                bank_transaction_id=bank_tx.id
            )

            # Update bank transaction category for quick access
            crud.update_bank_transaction_category(db, bank_tx.id, category)

            # Mark as approved if high confidence
            if auto_approved:
                crud.update_categorization_approval(
                    db, db_categorization.id, current_user.id, approved=True
                )
                high_confidence += 1
            else:
                low_confidence += 1

            # Track category distribution
            category_counts[category] = category_counts.get(category, 0) + 1

            results.append({
                "bank_transaction_id": bank_tx.id,
                "description": bank_tx.description,
                "amount": float(bank_tx.amount),
                "date": str(bank_tx.transaction_date) if bank_tx.transaction_date else None,
                "category": category,
                "subcategory": subcategory,
                "ledger_type": ledger_type,
                "confidence": confidence,
                "method": method,
                "explanation": explanation,
                "status": "categorized",
                "user_approved": auto_approved
            })
            processed += 1

        except Exception as e:
            print(f"Error categorizing transaction {bank_tx.id}: {str(e)}")
            results.append({
                "bank_transaction_id": bank_tx.id,
                "description": bank_tx.description,
                "amount": float(bank_tx.amount),
                "date": str(bank_tx.transaction_date) if bank_tx.transaction_date else None,
                "status": "error",
                "error": str(e)
            })
            failed += 1

    # Log activity
    crud.log_activity(
        db=db,
        user_id=current_user.id,
        action="batch_categorization",
        entity_type="bank_statement",
        entity_id=request.bank_statement_id,
        details={
            "total": len(bank_transactions),
            "processed": processed,
            "failed": failed,
            "high_confidence": high_confidence,
            "low_confidence": low_confidence
        }
    )

    return BatchCategorizationResponse(
        success=True,
        statement_id=request.bank_statement_id,
        total_transactions=len(bank_transactions),
        processed=processed,
        failed=failed,
        high_confidence=high_confidence,
        low_confidence=low_confidence,
        results=results,
        summary={
            "category_distribution": category_counts,
            "average_confidence": sum(r.get("confidence", 0) for r in results if r.get("status") != "error") / max(processed, 1),
            "needs_review_count": low_confidence,
            "auto_approved_count": high_confidence,
            "settings_applied": {
                "confidence_threshold": confidence_threshold,
                "vendor_mapping_enabled": use_vendor_mapping
            }
        }
    )


# ============================================================================
# ASYNC BATCH CATEGORIZATION WITH PROGRESS TRACKING
# ============================================================================

class AsyncBatchRequest(BaseModel):
    """Request model for async batch categorization"""
    bank_statement_id: int
    confidence_threshold: Optional[float] = Field(
        default=None, ge=0, le=100,
        description="If not provided, uses user's saved setting"
    )
    use_vendor_mapping: Optional[bool] = Field(
        default=None,
        description="If not provided, uses user's saved setting"
    )


class AsyncBatchResponse(BaseModel):
    """Response when starting an async batch job"""
    job_id: str
    status: str
    message: str
    total_transactions: int
    status_url: str


class JobStatusResponse(BaseModel):
    """Response model for job status"""
    job_id: str
    status: str
    progress_percent: float
    total_transactions: int
    processed_count: int
    failed_count: int
    high_confidence_count: int
    low_confidence_count: int
    current_transaction: str
    started_at: datetime
    completed_at: Optional[datetime]
    error_message: Optional[str]
    # Only included when completed
    results: Optional[list] = None
    summary: Optional[dict] = None


def process_batch_categorization_job(
    job_id: str,
    user_id: int,
    statement_id: int,
    confidence_threshold: float,
    use_vendor_mapping: bool,
    total_transactions: int,
    db_session_factory
):
    """Background task to process batch categorization with progress tracking"""
    from database import SessionLocal

    print(f"[BATCH] Starting batch job {job_id} for statement {statement_id} with {total_transactions} transactions")

    # Create a new database session for this background task
    db = SessionLocal()

    try:
        # Re-query the transactions in this session to avoid detached object issues
        bank_transactions = crud.get_bank_transactions_by_statement(db, user_id, statement_id)
        if not bank_transactions:
            print(f"[BATCH ERROR] No transactions found for statement {statement_id}")
            batch_job_tracker.fail_job(job_id, "No transactions found in database")
            return

        print(f"[BATCH] Retrieved {len(bank_transactions)} transactions from database")

        batch_job_tracker.start_job(job_id)
        print(f"[BATCH] Job {job_id} started, processing...")

        processed = 0
        failed = 0
        high_confidence = 0
        low_confidence = 0

        for bank_tx in bank_transactions:
            try:
                # Update progress
                batch_job_tracker.update_progress(
                    job_id, processed, bank_tx.description,
                    high_confidence, low_confidence, failed
                )

                # Check if already categorized
                existing_cat = crud.get_categorization_for_bank_transaction(
                    db, user_id, bank_tx.id
                )

                if existing_cat:
                    result = {
                        "bank_transaction_id": bank_tx.id,
                        "description": bank_tx.description,
                        "amount": float(bank_tx.amount),
                        "date": str(bank_tx.transaction_date) if bank_tx.transaction_date else None,
                        "category": existing_cat.category,
                        "subcategory": existing_cat.subcategory,
                        "ledger_type": existing_cat.ledger_type,
                        "confidence": float(existing_cat.confidence_score) if existing_cat.confidence_score else 0,
                        "method": existing_cat.method,
                        "status": "already_categorized",
                        "user_approved": existing_cat.user_approved
                    }
                    batch_job_tracker.add_result(job_id, result)
                    processed += 1
                    if existing_cat.confidence_score and existing_cat.confidence_score >= confidence_threshold:
                        high_confidence += 1
                    else:
                        low_confidence += 1
                    batch_job_tracker.update_category_count(job_id, existing_cat.category)
                    continue

                # Build document data structure
                document_data = {
                    "documentMetadata": {
                        "source": {"name": bank_tx.description},
                        "documentType": "bank_transaction",
                        "documentDate": str(bank_tx.transaction_date) if bank_tx.transaction_date else None,
                        "documentNumber": f"bank_tx_{bank_tx.id}"
                    },
                    "financialData": {
                        "totalAmount": float(bank_tx.amount),
                        "currency": "USD",
                        "transactionType": bank_tx.transaction_type or ("debit" if bank_tx.amount < 0 else "credit")
                    }
                }

                # Try vendor mapping first if enabled
                vendor_result = None
                if use_vendor_mapping:
                    vendor_result = categorize_by_vendor(bank_tx.description)

                if vendor_result:
                    category = vendor_result["category"]
                    subcategory = vendor_result["subcategory"]
                    ledger_type = vendor_result["ledger_type"]
                    confidence = vendor_result["confidence"]
                    method = "vendor_mapping"
                    explanation = vendor_result["explanation"]
                    ml_conf = 0
                    gemini_conf = 0
                else:
                    # Fall back to ML + Gemini
                    try:
                        engine = get_ml_categorization_engine()
                        # predict_category is async, need to run it with asyncio
                        ml_prediction = asyncio.run(engine.predict_category(document_data, "Bank statement transaction"))

                        if ml_prediction and ml_prediction.get("confidence", 0) > 50:
                            category = ml_prediction.get("category", "Operating Expenses")
                            subcategory = ml_prediction.get("subcategory", "General Operating")
                            ledger_type = ml_prediction.get("ledger_type", "Expense (Operating)")
                            confidence = ml_prediction.get("confidence", 50)
                            method = "hybrid"
                            explanation = ml_prediction.get("explanation", "ML prediction")
                            ml_conf = confidence
                            gemini_conf = 0
                        else:
                            # Use Gemini as fallback - also async
                            gemini_result = asyncio.run(_get_gemini_categorization(
                                bank_tx.description,
                                document_data,
                                "Bank statement transaction"
                            ))
                            category = gemini_result.get("category", "Operating Expenses")
                            subcategory = gemini_result.get("subcategory", "General Operating")
                            ledger_type = gemini_result.get("ledgerType", "Expense (Operating)")
                            confidence = gemini_result.get("confidence", 70)
                            method = "gemini"
                            explanation = gemini_result.get("explanation", "AI categorization")
                            ml_conf = 0
                            gemini_conf = confidence
                    except Exception as cat_error:
                        # If categorization fails, use defaults - mark as manual for review
                        category = "Operating Expenses"
                        subcategory = "General Operating"
                        ledger_type = "Expense (Operating)"
                        confidence = 30  # Low confidence since auto-categorization failed
                        method = "manual"  # Needs manual review
                        explanation = f"Categorization failed, needs manual review (error: {str(cat_error)[:50]})"
                        ml_conf = 0
                        gemini_conf = 0

                # Determine if auto-approved
                auto_approved = confidence >= confidence_threshold

                # Create categorization record
                categorization_data = {
                    "category": category,
                    "subcategory": subcategory,
                    "ledger_type": ledger_type,
                    "method": method,
                    "confidence_score": confidence,
                    "ml_confidence": ml_conf,
                    "gemini_confidence": gemini_conf,
                    "explanation": explanation,
                    "transaction_purpose": "Bank statement transaction"
                }

                db_categorization = crud.create_categorization(
                    db=db,
                    user_id=user_id,
                    categorization_data=categorization_data,
                    bank_transaction_id=bank_tx.id
                )

                # Update bank transaction category
                crud.update_bank_transaction_category(db, bank_tx.id, category)

                result = {
                    "bank_transaction_id": bank_tx.id,
                    "description": bank_tx.description,
                    "amount": float(bank_tx.amount),
                    "date": str(bank_tx.transaction_date) if bank_tx.transaction_date else None,
                    "category": category,
                    "subcategory": subcategory,
                    "ledger_type": ledger_type,
                    "confidence": confidence,
                    "method": method,
                    "explanation": explanation,
                    "status": "categorized",
                    "user_approved": auto_approved
                }

                batch_job_tracker.add_result(job_id, result)
                batch_job_tracker.update_category_count(job_id, category)

                processed += 1
                if confidence >= confidence_threshold:
                    high_confidence += 1
                else:
                    low_confidence += 1

            except Exception as tx_error:
                # Log the error for debugging
                print(f"[BATCH ERROR] Failed to categorize transaction {bank_tx.id}: {str(tx_error)}")
                import traceback
                traceback.print_exc()

                # Rollback the failed transaction to allow subsequent operations
                db.rollback()
                failed += 1
                batch_job_tracker.add_result(job_id, {
                    "bank_transaction_id": bank_tx.id,
                    "description": bank_tx.description,
                    "amount": float(bank_tx.amount),
                    "status": "error",
                    "error": str(tx_error)
                })

        # Update final progress
        batch_job_tracker.update_progress(
            job_id, processed, "", high_confidence, low_confidence, failed
        )

        # Log activity - use try-except to handle any remaining transaction issues
        try:
            # Ensure we have a clean transaction state
            db.rollback()
            crud.log_activity(
                db=db,
                user_id=user_id,
                action="async_batch_categorization",
                entity_type="bank_statement",
                entity_id=statement_id,
                details={
                    "job_id": job_id,
                    "total": len(bank_transactions),
                    "processed": processed,
                    "failed": failed
                }
            )
            db.commit()
        except Exception as log_error:
            print(f"Error logging activity: {log_error}")
            db.rollback()

        # Mark job as completed
        batch_job_tracker.complete_job(job_id, success=True)

    except Exception as e:
        batch_job_tracker.complete_job(job_id, success=False, error_message=str(e))

    finally:
        db.close()


@app.post("/categorize-bank-statement/async", response_model=AsyncBatchResponse, tags=["Batch Processing"])
async def start_async_batch_categorization(
    request: AsyncBatchRequest,
    background_tasks: BackgroundTasks,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Start async batch categorization with progress tracking.

    Returns immediately with a job_id that can be used to check progress.
    Use GET /batch-job/{job_id} to check status and get results.

    This is ideal for large statements where you want to show a progress bar.
    """
    # Verify the bank statement exists
    statement = crud.get_bank_statement_by_id(db, current_user.id, request.bank_statement_id)
    if not statement:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Bank statement with ID {request.bank_statement_id} not found"
        )

    # Get transactions
    bank_transactions = crud.get_bank_transactions_by_statement(
        db, current_user.id, request.bank_statement_id
    )

    if not bank_transactions:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No transactions found in this statement"
        )

    # Resolve settings
    user_settings = current_user.settings or {}
    confidence_threshold = request.confidence_threshold
    if confidence_threshold is None:
        confidence_threshold = user_settings.get("confidence_threshold", DEFAULT_USER_SETTINGS["confidence_threshold"])

    use_vendor_mapping = request.use_vendor_mapping
    if use_vendor_mapping is None:
        use_vendor_mapping = user_settings.get("auto_approve_vendor_mapping", DEFAULT_USER_SETTINGS["auto_approve_vendor_mapping"])

    # Create job
    job_id = batch_job_tracker.create_job(
        user_id=current_user.id,
        statement_id=request.bank_statement_id,
        total_transactions=len(bank_transactions)
    )

    # Need to import SessionLocal for the background task
    from database import SessionLocal

    # Start background task - pass transaction count instead of ORM objects
    # to avoid detached object issues in the background thread
    background_tasks.add_task(
        process_batch_categorization_job,
        job_id=job_id,
        user_id=current_user.id,
        statement_id=request.bank_statement_id,
        confidence_threshold=confidence_threshold,
        use_vendor_mapping=use_vendor_mapping,
        total_transactions=len(bank_transactions),
        db_session_factory=SessionLocal
    )

    return AsyncBatchResponse(
        job_id=job_id,
        status="pending",
        message=f"Batch categorization started for {len(bank_transactions)} transactions",
        total_transactions=len(bank_transactions),
        status_url=f"/batch-job/{job_id}"
    )


@app.get("/batch-job/{job_id}", response_model=JobStatusResponse, tags=["Batch Processing"])
async def get_batch_job_status(
    job_id: str,
    current_user: models.User = Depends(get_current_user)
):
    """
    Get the status and progress of a batch categorization job.

    Poll this endpoint to track progress. When status is 'completed',
    the results and summary will be included in the response.
    """
    job = batch_job_tracker.get_job(job_id)

    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Job {job_id} not found"
        )

    # Verify job belongs to user
    if job.user_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this job"
        )

    response = JobStatusResponse(
        job_id=job.job_id,
        status=job.status,
        progress_percent=job.progress_percent,
        total_transactions=job.total_transactions,
        processed_count=job.processed_count,
        failed_count=job.failed_count,
        high_confidence_count=job.high_confidence_count,
        low_confidence_count=job.low_confidence_count,
        current_transaction=job.current_transaction,
        started_at=job.started_at,
        completed_at=job.completed_at,
        error_message=job.error_message
    )

    # Include results and summary only when completed
    if job.status == "completed":
        response.results = job.results
        response.summary = {
            "category_distribution": job.category_counts,
            "average_confidence": sum(r.get("confidence", 0) for r in job.results if r.get("status") != "error") / max(job.processed_count, 1),
            "needs_review_count": job.low_confidence_count,
            "auto_approved_count": job.high_confidence_count
        }

    return response


@app.get("/batch-jobs", tags=["Batch Processing"])
async def list_batch_jobs(
    limit: int = Query(default=10, ge=1, le=50),
    current_user: models.User = Depends(get_current_user)
):
    """
    List recent batch jobs for the current user.

    Returns job status summaries (not full results).
    """
    jobs = batch_job_tracker.get_user_jobs(current_user.id, limit)

    return {
        "jobs": [
            {
                "job_id": j.job_id,
                "statement_id": j.statement_id,
                "status": j.status,
                "progress_percent": j.progress_percent,
                "total_transactions": j.total_transactions,
                "processed_count": j.processed_count,
                "started_at": j.started_at,
                "completed_at": j.completed_at
            }
            for j in jobs
        ],
        "count": len(jobs)
    }


# ============================================================================
# EXPORT ENDPOINTS
# ============================================================================

class ExportFilter(str, Enum):
    """Filter options for export endpoints"""
    all = "all"
    approved = "approved"
    needs_review = "needs_review"
    uncategorized = "uncategorized"
    high_confidence = "high_confidence"
    low_confidence = "low_confidence"


@app.get("/export-statement/{statement_id}", tags=["Export"])
async def export_statement_csv(
    statement_id: int,
    filter: ExportFilter = Query(default=ExportFilter.all, description="Filter transactions"),
    confidence_threshold: Optional[float] = Query(default=None, description="Confidence threshold for high/low filter"),
    date_start: Optional[date] = Query(default=None, description="Filter transactions from this date (YYYY-MM-DD)"),
    date_end: Optional[date] = Query(default=None, description="Filter transactions until this date (YYYY-MM-DD)"),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Export categorized bank statement transactions to CSV format.

    **Filter options:**
    - `all`: Export all transactions
    - `approved`: Only export approved transactions
    - `needs_review`: Only export transactions needing review (low confidence or not approved)
    - `uncategorized`: Only export uncategorized transactions
    - `high_confidence`: Only export high confidence transactions (>= threshold)
    - `low_confidence`: Only export low confidence transactions (< threshold)

    **Date range filtering:**
    - `date_start`: Only include transactions on or after this date
    - `date_end`: Only include transactions on or before this date

    Returns a CSV file with columns:
    - Date, Description, Amount, Type, Category, Subcategory, Ledger Type, Confidence, Method, Approved
    """
    from fastapi.responses import StreamingResponse
    import csv

    # Get user's confidence threshold if not provided
    if confidence_threshold is None:
        user_settings = current_user.settings or {}
        confidence_threshold = user_settings.get("confidence_threshold", 70.0)

    # Verify the bank statement exists and belongs to the user
    statement = crud.get_bank_statement_by_id(db, current_user.id, statement_id)
    if not statement:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Bank statement with ID {statement_id} not found"
        )

    # Get all transactions with their categorizations
    bank_transactions = crud.get_bank_transactions_by_statement(
        db, current_user.id, statement_id
    )

    # Apply date range filter if provided
    if date_start or date_end:
        date_filtered = []
        for tx in bank_transactions:
            tx_date = tx.transaction_date
            if tx_date:
                # Convert to date if it's datetime
                if hasattr(tx_date, 'date'):
                    tx_date = tx_date.date()
                elif isinstance(tx_date, str):
                    tx_date = datetime.strptime(tx_date[:10], "%Y-%m-%d").date()

                if date_start and tx_date < date_start:
                    continue
                if date_end and tx_date > date_end:
                    continue
            date_filtered.append(tx)
        bank_transactions = date_filtered

    # Build list of (transaction, categorization) tuples with filtering
    filtered_transactions = []
    for tx in bank_transactions:
        categorization = crud.get_categorization_for_bank_transaction(
            db, current_user.id, tx.id
        )

        # Apply filter
        include = False
        if filter == ExportFilter.all:
            include = True
        elif filter == ExportFilter.approved:
            include = categorization and categorization.user_approved
        elif filter == ExportFilter.needs_review:
            include = not categorization or not categorization.user_approved or (
                categorization.confidence_score and categorization.confidence_score < confidence_threshold
            )
        elif filter == ExportFilter.uncategorized:
            include = not categorization
        elif filter == ExportFilter.high_confidence:
            include = categorization and categorization.confidence_score and categorization.confidence_score >= confidence_threshold
        elif filter == ExportFilter.low_confidence:
            include = categorization and categorization.confidence_score and categorization.confidence_score < confidence_threshold

        if include:
            filtered_transactions.append((tx, categorization))

    # Build CSV data
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow([
        "Date",
        "Description",
        "Amount",
        "Type",
        "Category",
        "Subcategory",
        "Ledger Type",
        "Confidence",
        "Method",
        "Approved"
    ])

    # Write data rows
    for tx, categorization in filtered_transactions:
        tx_type = tx.transaction_type or ("debit" if tx.amount and tx.amount < 0 else "credit")

        writer.writerow([
            str(tx.transaction_date) if tx.transaction_date else "",
            tx.description or "",
            float(tx.amount) if tx.amount else 0,
            tx_type,
            categorization.category if categorization else "",
            categorization.subcategory if categorization else "",
            categorization.ledger_type if categorization else "",
            float(categorization.confidence_score) if categorization and categorization.confidence_score else "",
            categorization.method if categorization else "",
            "Yes" if categorization and categorization.user_approved else "No"
        ])

    # Prepare response
    output.seek(0)

    # Generate filename with filter info
    filter_suffix = f"_{filter.value}" if filter != ExportFilter.all else ""
    filename = f"categorized_statement_{statement_id}{filter_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    # Log activity
    crud.log_activity(
        db=db,
        user_id=current_user.id,
        action="export_csv",
        entity_type="bank_statement",
        entity_id=statement_id,
        details={
            "transaction_count": len(filtered_transactions),
            "total_transactions": len(bank_transactions),
            "filter": filter.value
        }
    )

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/export-statement/{statement_id}/excel", tags=["Export"])
async def export_statement_excel(
    statement_id: int,
    filter: ExportFilter = Query(default=ExportFilter.all, description="Filter transactions"),
    confidence_threshold: Optional[float] = Query(default=None, description="Confidence threshold for high/low filter"),
    date_start: Optional[date] = Query(default=None, description="Filter transactions from this date (YYYY-MM-DD)"),
    date_end: Optional[date] = Query(default=None, description="Filter transactions until this date (YYYY-MM-DD)"),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Export categorized bank statement transactions to Excel format.

    **Filter options:**
    - `all`: Export all transactions
    - `approved`: Only export approved transactions
    - `needs_review`: Only export transactions needing review
    - `uncategorized`: Only export uncategorized transactions
    - `high_confidence`: Only export high confidence transactions
    - `low_confidence`: Only export low confidence transactions

    **Date range filtering:**
    - `date_start`: Only include transactions on or after this date
    - `date_end`: Only include transactions on or before this date

    Returns an Excel file with:
    - Summary sheet with statistics
    - Detail sheet with filtered transactions and categorizations
    """
    from fastapi.responses import StreamingResponse

    # Get user's confidence threshold if not provided
    if confidence_threshold is None:
        user_settings = current_user.settings or {}
        confidence_threshold = user_settings.get("confidence_threshold", 70.0)

    # Verify the bank statement exists and belongs to the user
    statement = crud.get_bank_statement_by_id(db, current_user.id, statement_id)
    if not statement:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Bank statement with ID {statement_id} not found"
        )

    # Get all transactions with their categorizations
    bank_transactions = crud.get_bank_transactions_by_statement(
        db, current_user.id, statement_id
    )

    # Apply date range filter if provided
    if date_start or date_end:
        date_filtered = []
        for tx in bank_transactions:
            tx_date = tx.transaction_date
            if tx_date:
                # Convert to date if it's datetime
                if hasattr(tx_date, 'date'):
                    tx_date = tx_date.date()
                elif isinstance(tx_date, str):
                    tx_date = datetime.strptime(tx_date[:10], "%Y-%m-%d").date()

                if date_start and tx_date < date_start:
                    continue
                if date_end and tx_date > date_end:
                    continue
            date_filtered.append(tx)
        bank_transactions = date_filtered

    # Build list of (transaction, categorization) tuples with filtering
    filtered_transactions = []
    for tx in bank_transactions:
        categorization = crud.get_categorization_for_bank_transaction(
            db, current_user.id, tx.id
        )

        # Apply filter
        include = False
        if filter == ExportFilter.all:
            include = True
        elif filter == ExportFilter.approved:
            include = categorization and categorization.user_approved
        elif filter == ExportFilter.needs_review:
            include = not categorization or not categorization.user_approved or (
                categorization.confidence_score and categorization.confidence_score < confidence_threshold
            )
        elif filter == ExportFilter.uncategorized:
            include = not categorization
        elif filter == ExportFilter.high_confidence:
            include = categorization and categorization.confidence_score and categorization.confidence_score >= confidence_threshold
        elif filter == ExportFilter.low_confidence:
            include = categorization and categorization.confidence_score and categorization.confidence_score < confidence_threshold

        if include:
            filtered_transactions.append((tx, categorization))

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        # Create workbook
        wb = openpyxl.Workbook()

        # Summary sheet
        summary_sheet = wb.active
        summary_sheet.title = "Summary"

        # Calculate statistics from filtered transactions
        total_transactions = len(bank_transactions)
        filtered_count = len(filtered_transactions)
        categorized_count = 0
        approved_count = 0
        total_amount = 0
        category_totals = {}

        for tx, categorization in filtered_transactions:
            total_amount += float(tx.amount) if tx.amount else 0
            if categorization:
                categorized_count += 1
                if categorization.user_approved:
                    approved_count += 1
                cat = categorization.category
                if cat not in category_totals:
                    category_totals[cat] = {"count": 0, "amount": 0}
                category_totals[cat]["count"] += 1
                category_totals[cat]["amount"] += float(tx.amount) if tx.amount else 0

        # Write summary
        summary_sheet["A1"] = "Bank Statement Export Summary"
        summary_sheet["A1"].font = Font(bold=True, size=14)

        filter_display = filter.value.replace("_", " ").title()
        summary_data = [
            ("Statement ID", statement_id),
            ("File Name", statement.file_name),
            ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Filter Applied", filter_display),
            ("", ""),
            ("Total in Statement", total_transactions),
            ("Exported Transactions", filtered_count),
            ("Categorized", categorized_count),
            ("Approved", approved_count),
            ("Total Amount", f"${total_amount:,.2f}"),
            ("", ""),
            ("Category Breakdown", ""),
        ]

        row = 3
        for label, value in summary_data:
            summary_sheet[f"A{row}"] = label
            summary_sheet[f"B{row}"] = value
            if label:
                summary_sheet[f"A{row}"].font = Font(bold=True)
            row += 1

        # Category breakdown
        for cat, data in sorted(category_totals.items()):
            summary_sheet[f"A{row}"] = f"  {cat}"
            summary_sheet[f"B{row}"] = f"{data['count']} transactions (${data['amount']:,.2f})"
            row += 1

        # Detail sheet
        detail_sheet = wb.create_sheet("Transactions")

        # Headers
        headers = [
            "Date", "Description", "Amount", "Type",
            "Category", "Subcategory", "Ledger Type",
            "Confidence", "Method", "Approved"
        ]

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = detail_sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows - use filtered transactions
        for row_num, (tx, categorization) in enumerate(filtered_transactions, 2):

            tx_type = tx.transaction_type or ("debit" if tx.amount and tx.amount < 0 else "credit")

            detail_sheet.cell(row=row_num, column=1, value=str(tx.transaction_date) if tx.transaction_date else "")
            detail_sheet.cell(row=row_num, column=2, value=tx.description or "")
            detail_sheet.cell(row=row_num, column=3, value=float(tx.amount) if tx.amount else 0)
            detail_sheet.cell(row=row_num, column=4, value=tx_type)
            detail_sheet.cell(row=row_num, column=5, value=categorization.category if categorization else "")
            detail_sheet.cell(row=row_num, column=6, value=categorization.subcategory if categorization else "")
            detail_sheet.cell(row=row_num, column=7, value=categorization.ledger_type if categorization else "")
            detail_sheet.cell(row=row_num, column=8, value=float(categorization.confidence_score) if categorization and categorization.confidence_score else "")
            detail_sheet.cell(row=row_num, column=9, value=categorization.method if categorization else "")
            detail_sheet.cell(row=row_num, column=10, value="Yes" if categorization and categorization.user_approved else "No")

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            detail_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        detail_sheet.column_dimensions["B"].width = 40  # Description column wider

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with filter info
        filter_suffix = f"_{filter.value}" if filter != ExportFilter.all else ""
        filename = f"categorized_statement_{statement_id}{filter_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Log activity
        crud.log_activity(
            db=db,
            user_id=current_user.id,
            action="export_excel",
            entity_type="bank_statement",
            entity_id=statement_id,
            details={
                "transaction_count": len(filtered_transactions),
                "total_transactions": len(bank_transactions),
                "filter": filter.value
            }
        )

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except ImportError:
        # openpyxl not installed, return error with instructions
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Excel export requires openpyxl library. Install with: pip install openpyxl"
        )


@app.get("/export-statement/{statement_id}/quickbooks", tags=["Export"])
async def export_statement_quickbooks(
    statement_id: int,
    filter: ExportFilter = Query(default=ExportFilter.approved, description="Filter transactions (default: approved only)"),
    confidence_threshold: Optional[float] = Query(default=None, description="Confidence threshold"),
    date_start: Optional[date] = Query(default=None, description="Filter transactions from this date (YYYY-MM-DD)"),
    date_end: Optional[date] = Query(default=None, description="Filter transactions until this date (YYYY-MM-DD)"),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Export categorized bank statement in QuickBooks-compatible CSV format.

    **Format:** CSV compatible with QuickBooks Online bank transaction import.

    **Columns:**
    - Date (MM/DD/YYYY format)
    - Description
    - Amount (positive for deposits, negative for withdrawals)
    - Category (mapped to QuickBooks account name)

    **Date range filtering:**
    - `date_start`: Only include transactions on or after this date
    - `date_end`: Only include transactions on or before this date

    **Note:** By default, only exports approved transactions to ensure data quality.
    """
    from fastapi.responses import StreamingResponse
    import csv

    # Get user's confidence threshold if not provided
    if confidence_threshold is None:
        user_settings = current_user.settings or {}
        confidence_threshold = user_settings.get("confidence_threshold", 70.0)

    # Verify the bank statement exists and belongs to the user
    statement = crud.get_bank_statement_by_id(db, current_user.id, statement_id)
    if not statement:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Bank statement with ID {statement_id} not found"
        )

    # Get all transactions
    bank_transactions = crud.get_bank_transactions_by_statement(
        db, current_user.id, statement_id
    )

    # Apply date range filter if provided
    if date_start or date_end:
        date_filtered = []
        for tx in bank_transactions:
            tx_date = tx.transaction_date
            if tx_date:
                # Convert to date if it's datetime
                if hasattr(tx_date, 'date'):
                    tx_date = tx_date.date()
                elif isinstance(tx_date, str):
                    tx_date = datetime.strptime(tx_date[:10], "%Y-%m-%d").date()

                if date_start and tx_date < date_start:
                    continue
                if date_end and tx_date > date_end:
                    continue
            date_filtered.append(tx)
        bank_transactions = date_filtered

    # Build filtered list
    filtered_transactions = []
    for tx in bank_transactions:
        categorization = crud.get_categorization_for_bank_transaction(
            db, current_user.id, tx.id
        )

        # Apply filter
        include = False
        if filter == ExportFilter.all:
            include = True
        elif filter == ExportFilter.approved:
            include = categorization and categorization.user_approved
        elif filter == ExportFilter.needs_review:
            include = not categorization or not categorization.user_approved or (
                categorization.confidence_score and categorization.confidence_score < confidence_threshold
            )
        elif filter == ExportFilter.uncategorized:
            include = not categorization
        elif filter == ExportFilter.high_confidence:
            include = categorization and categorization.confidence_score and categorization.confidence_score >= confidence_threshold
        elif filter == ExportFilter.low_confidence:
            include = categorization and categorization.confidence_score and categorization.confidence_score < confidence_threshold

        if include:
            filtered_transactions.append((tx, categorization))

    # Build QuickBooks-compatible CSV
    output = io.StringIO()
    writer = csv.writer(output)

    # QuickBooks CSV header
    writer.writerow(["Date", "Description", "Amount", "Category"])

    # Write data rows in QuickBooks format
    for tx, categorization in filtered_transactions:
        # Format date as MM/DD/YYYY for QuickBooks
        date_str = ""
        if tx.transaction_date:
            try:
                if isinstance(tx.transaction_date, str):
                    # Parse and reformat
                    from datetime import datetime as dt
                    parsed = dt.strptime(str(tx.transaction_date)[:10], "%Y-%m-%d")
                    date_str = parsed.strftime("%m/%d/%Y")
                else:
                    date_str = tx.transaction_date.strftime("%m/%d/%Y")
            except:
                date_str = str(tx.transaction_date)

        # Build category path for QuickBooks (Category:Subcategory format)
        category_str = ""
        if categorization:
            if categorization.category and categorization.subcategory:
                category_str = f"{categorization.category}:{categorization.subcategory}"
            elif categorization.category:
                category_str = categorization.category

        writer.writerow([
            date_str,
            tx.description or "",
            float(tx.amount) if tx.amount else 0,
            category_str
        ])

    output.seek(0)

    # Generate filename
    filter_suffix = f"_{filter.value}" if filter != ExportFilter.approved else ""
    filename = f"quickbooks_import_{statement_id}{filter_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    # Log activity
    crud.log_activity(
        db=db,
        user_id=current_user.id,
        action="export_quickbooks",
        entity_type="bank_statement",
        entity_id=statement_id,
        details={
            "transaction_count": len(filtered_transactions),
            "total_transactions": len(bank_transactions),
            "filter": filter.value
        }
    )

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/bank-statement/{statement_id}/status")
async def get_statement_categorization_status(
    statement_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Get the categorization status for a bank statement.

    Returns statistics about how many transactions are categorized,
    approved, and need review.
    """
    # Verify the bank statement exists and belongs to the user
    statement = crud.get_bank_statement_by_id(db, current_user.id, statement_id)
    if not statement:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Bank statement with ID {statement_id} not found"
        )

    # Get all transactions
    bank_transactions = crud.get_bank_transactions_by_statement(
        db, current_user.id, statement_id
    )

    total = len(bank_transactions)
    categorized = 0
    approved = 0
    needs_review = 0
    uncategorized = 0

    for tx in bank_transactions:
        categorization = crud.get_categorization_for_bank_transaction(
            db, current_user.id, tx.id
        )
        if categorization:
            categorized += 1
            if categorization.user_approved:
                approved += 1
            else:
                needs_review += 1
        else:
            uncategorized += 1

    return {
        "statement_id": statement_id,
        "file_name": statement.file_name,
        "total_transactions": total,
        "categorized": categorized,
        "approved": approved,
        "needs_review": needs_review,
        "uncategorized": uncategorized,
        "completion_percentage": (categorized / total * 100) if total > 0 else 0
    }


@app.get("/bank-statement/{statement_id}/results", tags=["Bank Statements"])
async def get_statement_results(
    statement_id: int,
    sort_by: Optional[str] = Query(default="date", description="Sort by: date, amount, category, confidence, status"),
    sort_order: Optional[str] = Query(default="desc", description="Sort order: asc or desc"),
    filter_status: Optional[str] = Query(default=None, description="Filter by status: approved, needs_review, uncategorized"),
    filter_category: Optional[str] = Query(default=None, description="Filter by category name"),
    min_confidence: Optional[float] = Query(default=None, description="Minimum confidence score (0-100)"),
    max_confidence: Optional[float] = Query(default=None, description="Maximum confidence score (0-100)"),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Get complete categorization results for a bank statement.

    Returns all transactions with their categorizations, summary statistics,
    and category breakdown. Supports sorting and filtering.

    **Sort options:** date, amount, category, confidence, status
    **Filter options:** approved, needs_review, uncategorized
    """
    # Verify the bank statement exists and belongs to the user
    statement = crud.get_bank_statement_by_id(db, current_user.id, statement_id)
    if not statement:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Bank statement with ID {statement_id} not found"
        )

    # Get user's confidence threshold
    user_settings = current_user.settings or {}
    confidence_threshold = user_settings.get("confidence_threshold", 70.0)

    # Get all transactions
    bank_transactions = crud.get_bank_transactions_by_statement(
        db, current_user.id, statement_id
    )

    # Build results with categorizations
    transactions = []
    stats = {
        "total": 0,
        "categorized": 0,
        "approved": 0,
        "needs_review": 0,
        "uncategorized": 0,
        "high_confidence": 0,
        "low_confidence": 0,
        "total_amount": 0,
        "approved_amount": 0
    }
    category_breakdown = {}

    for tx in bank_transactions:
        categorization = crud.get_categorization_for_bank_transaction(
            db, current_user.id, tx.id
        )

        # Determine status
        if categorization:
            if categorization.user_approved:
                tx_status = "approved"
                stats["approved"] += 1
                stats["approved_amount"] += float(tx.amount) if tx.amount else 0
            else:
                tx_status = "needs_review"
                stats["needs_review"] += 1
            stats["categorized"] += 1

            # Confidence tracking
            conf = categorization.confidence_score or 0
            if conf >= confidence_threshold:
                stats["high_confidence"] += 1
            else:
                stats["low_confidence"] += 1

            # Category breakdown
            cat = categorization.category or "Uncategorized"
            if cat not in category_breakdown:
                category_breakdown[cat] = {"count": 0, "amount": 0, "approved": 0}
            category_breakdown[cat]["count"] += 1
            category_breakdown[cat]["amount"] += float(tx.amount) if tx.amount else 0
            if categorization.user_approved:
                category_breakdown[cat]["approved"] += 1
        else:
            tx_status = "uncategorized"
            stats["uncategorized"] += 1
            cat = "Uncategorized"
            if cat not in category_breakdown:
                category_breakdown[cat] = {"count": 0, "amount": 0, "approved": 0}
            category_breakdown[cat]["count"] += 1
            category_breakdown[cat]["amount"] += float(tx.amount) if tx.amount else 0

        stats["total"] += 1
        stats["total_amount"] += float(tx.amount) if tx.amount else 0

        # Apply filters
        if filter_status and tx_status != filter_status:
            continue
        if filter_category and categorization and categorization.category != filter_category:
            continue
        if filter_category and not categorization and filter_category != "Uncategorized":
            continue
        if min_confidence is not None:
            if not categorization or (categorization.confidence_score or 0) < min_confidence:
                continue
        if max_confidence is not None:
            if categorization and (categorization.confidence_score or 0) > max_confidence:
                continue

        # Build transaction record
        tx_record = {
            "id": tx.id,
            "transaction_date": str(tx.transaction_date) if tx.transaction_date else None,
            "description": tx.description,
            "amount": float(tx.amount) if tx.amount else 0,
            "transaction_type": tx.transaction_type or ("debit" if tx.amount and tx.amount < 0 else "credit"),
            "status": tx_status,
            "category": categorization.category if categorization else None,
            "subcategory": categorization.subcategory if categorization else None,
            "ledger_type": categorization.ledger_type if categorization else None,
            "confidence": float(categorization.confidence_score) if categorization and categorization.confidence_score else None,
            "method": categorization.method if categorization else None,
            "user_approved": categorization.user_approved if categorization else False,
            "categorization_id": categorization.id if categorization else None
        }
        transactions.append(tx_record)

    # Sort transactions
    sort_keys = {
        "date": lambda x: x["transaction_date"] or "",
        "amount": lambda x: abs(x["amount"]),
        "category": lambda x: x["category"] or "zzz",  # Put None at end
        "confidence": lambda x: x["confidence"] or 0,
        "status": lambda x: x["status"]
    }

    if sort_by in sort_keys:
        reverse = sort_order.lower() == "desc"
        transactions.sort(key=sort_keys[sort_by], reverse=reverse)

    # Calculate percentages
    stats["completion_percentage"] = (stats["categorized"] / stats["total"] * 100) if stats["total"] > 0 else 0
    stats["approval_percentage"] = (stats["approved"] / stats["categorized"] * 100) if stats["categorized"] > 0 else 0

    return {
        "statement": {
            "id": statement.id,
            "file_name": statement.file_name,
            "bank_name": statement.bank_name,
            "account_number": statement.account_number,
            "period_start": str(statement.period_start) if statement.period_start else None,
            "period_end": str(statement.period_end) if statement.period_end else None,
            "uploaded_at": str(statement.uploaded_at) if statement.uploaded_at else None
        },
        "summary": stats,
        "category_breakdown": [
            {"category": cat, **data}
            for cat, data in sorted(category_breakdown.items(), key=lambda x: x[1]["count"], reverse=True)
        ],
        "transactions": transactions,
        "filters_applied": {
            "sort_by": sort_by,
            "sort_order": sort_order,
            "filter_status": filter_status,
            "filter_category": filter_category,
            "min_confidence": min_confidence,
            "max_confidence": max_confidence
        },
        "confidence_threshold": confidence_threshold
    }


@app.get("/known-vendors")
async def list_known_vendors():
    """
    List all known vendor mappings used for deterministic categorization.

    These vendors are categorized instantly without AI calls,
    providing faster and more consistent results.
    """
    vendors = get_all_known_vendors()

    # Group by category for better organization
    by_category = {}
    for pattern, mapping in vendors.items():
        cat = mapping["category"]
        if cat not in by_category:
            by_category[cat] = []
        by_category[cat].append({
            "pattern": pattern,
            "subcategory": mapping["subcategory"],
            "ledger_type": mapping["ledger_type"]
        })

    return {
        "total_vendors": len(vendors),
        "vendors": vendors,
        "by_category": by_category
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)